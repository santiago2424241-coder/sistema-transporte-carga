"""
Sistema de Programación de Rutas y Cálculo de Costos para Tractomulas
Versión 4.8 - Conectado a Supabase (PostgreSQL) - ACTUALIZADO
Contexto: Colombia
Autor: Sistema de Gestión de Transporte de Carga

CAMBIOS EN ESTA VERSIÓN (v4.6 - FLUIDEZ DE NAVEGACIÓN Y LISTAS):
- CORREGIDO: la navegación por pestañas (radio superior) ahora usa
  on_change con un callback en vez de reasignar st.session_state.tab_actual
  en la misma línea donde se crea el widget. Esto evita que haya que hacer
  clic dos veces para que el contenido de la pestaña cambie.
- NUEVO: las listas de Tractomulas, Conductores y Rutas ahora se cargan con
  funciones envueltas en @st.cache_data (TTL 30s), y se invalida el cache
  explícitamente (.clear()) justo después de guardar/eliminar cualquiera de
  ellas. Antes se releían de la base de datos en cada rerun sin cache,
  sumando latencia de red a cada clic (incluido el propio cambio de pestaña).
- NUEVO: los combos "Filtrar / Buscar" de Trazabilidad (Tab 6), Acumulado
  por Flota (Tab 7) y Liquidaciones (Tab 8) ahora guardan y reutilizan el
  resultado en session_state de forma explícita en vez de mezclar la
  condición del botón con "not in session_state", evitando que el resultado
  visible quede un rerun "atrasado" respecto al filtro que se ve en pantalla.

CAMBIOS EN ESTA VERSIÓN (v4.7 - CONTEO REAL DE VIAJES POR numero_viajes):
- CORREGIDO: en Dashboard, Estadísticas Generales, Trazabilidad y
  Liquidaciones, el conteo de "cuántos viajes hizo" un conductor/tractomula/
  ruta ahora SUMA el campo numero_viajes de cada registro, en vez de contar
  filas (COUNT(*) / len(df)). Esto aplica para CUALQUIER cliente, no solo
  AGOFER: si registras un solo viaje con Número de Viajes = 3, ahora cuenta
  como 3 viajes en todos los reportes, no como 1.

CAMBIOS EN ESTA VERSIÓN (v4.8 - DÍAS SIN VIAJE):
- NUEVO: en la pestaña "4. Cálculo de Viaje" se agregó un checkbox
  "🚫 Día vacío (el carro NO hizo viaje este día)". Al activarlo se oculta
  todo el formulario de cálculo de costos y solo se guarda un registro
  simple (fecha, placa, conductor opcional, motivo, observaciones) en la
  nueva tabla `dias_sin_viaje`. NO se calcula ni se guarda ningún costo,
  gasto, flete ni utilidad para estos registros — son solo para
  trazabilidad/historial.
- NUEVO: en la pestaña "6. Trazabilidad" se agregó una sección para ver,
  filtrar por placa y eliminar estos registros de días sin viaje.
"""

import streamlit as st
import re
import psycopg2
from psycopg2 import sql
from psycopg2 import pool as pg_pool
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import List, Dict
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import locale
import plotly.express as px

# Configurar locale para formato colombiano
try:
    locale.setlocale(locale.LC_ALL, 'es_CO.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
    except:
        pass


# ==================== CONFIGURACIÓN SUPABASE ====================
SUPABASE_DB_URL = "postgresql://postgres.wiomyjrmsrhcgvhgkbqe:Conejito800$@aws-1-us-west-2.pooler.supabase.com:6543/postgres"


# ==================== POOL DE CONEXIONES ====================
@st.cache_resource
def get_db_pool():
    return pg_pool.ThreadedConnectionPool(1, 20, SUPABASE_DB_URL)


_db_initialized = False


# ==================== FUNCIONES DE FORMATO ====================
def formatear_numero(valor):
    """Formatea un número al estilo colombiano: 5.000.000"""
    if valor is None:
        return "0"
    try:
        return f"{int(valor):,}".replace(',', '.')
    except:
        return str(valor)


def formatear_decimal(valor, decimales=2):
    """Formatea un número con decimales al estilo colombiano: 5.000.000,50"""
    if valor is None:
        return "0,00"
    try:
        formatted = f"{float(valor):,.{decimales}f}"
        formatted = formatted.replace(',', 'TEMP')
        formatted = formatted.replace('.', ',')
        formatted = formatted.replace('TEMP', '.')
        return formatted
    except:
        return str(valor)


def limpiar_numero(texto):
    """Convierte texto con formato colombiano a número"""
    if not texto:
        return 0.0
    try:
        texto = str(texto).replace('.', '').replace(',', '.')
        return float(texto)
    except:
        return 0.0


def es_cliente_agofer(cliente: str) -> bool:
    """Normaliza el nombre del cliente para detectar si corresponde a 'AGOFER'."""
    if not cliente:
        return False
    return str(cliente).strip().upper() == "AGOFER"


# ==================== BASE DE DATOS SUPABASE ====================
class DatabaseManager:
    """Gestor de base de datos Supabase (PostgreSQL) para trazabilidad."""

    def __init__(self):
        self.pool = get_db_pool()
        self.init_database()

    def get_connection(self):
        return self.pool.getconn()

    def release_connection(self, conn):
        try:
            self.pool.putconn(conn)
        except Exception:
            try:
                conn.close()
            except Exception:
                pass

    def init_database(self):
        global _db_initialized
        if _db_initialized:
            return
        conn = None
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS viajes_v4 (
                    id SERIAL PRIMARY KEY,
                    fecha_creacion TEXT NOT NULL,
                    placa TEXT NOT NULL,
                    conductor TEXT NOT NULL,
                    origen TEXT NOT NULL,
                    destino TEXT NOT NULL,
                    distancia_km REAL NOT NULL,
                    dias_viaje INTEGER NOT NULL,
                    es_frontera INTEGER NOT NULL,
                    hubo_parqueo INTEGER NOT NULL,
                    nomina_admin REAL,
                    nomina_conductor REAL,
                    comision_conductor REAL,
                    mantenimiento REAL,
                    seguros REAL,
                    tecnomecanica REAL,
                    llantas REAL,
                    aceite REAL,
                    combustible REAL,
                    galones_necesarios REAL,
                    flypass REAL,
                    peajes REAL,
                    cruce_frontera REAL,
                    hotel REAL,
                    comida REAL,
                    parqueo REAL,
                    cargue_descargue REAL,
                    otros REAL,
                    total_gastos REAL,
                    legalizacion REAL,
                    punto_equilibrio REAL,
                    valor_flete REAL,
                    utilidad REAL,
                    rentabilidad REAL,
                    anticipo REAL,
                    saldo REAL,
                    hubo_anticipo_empresa INTEGER,
                    ant_empresa REAL,
                    saldo_empresa REAL,
                    observaciones TEXT
                )
            ''')

            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS urea_acpm REAL DEFAULT 0")
            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS transporte REAL DEFAULT 0")
            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS propina_comision REAL DEFAULT 0")

            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS fecha_viaje DATE")
            cursor.execute("""
                UPDATE viajes_v4 SET fecha_viaje = to_date(fecha_creacion, 'YYYY-MM-DD')
                WHERE fecha_viaje IS NULL
            """)

            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS cliente TEXT")
            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS galones_reales REAL")
            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS numero_viajes INTEGER DEFAULT 1")
            cursor.execute("ALTER TABLE viajes_v4 ADD COLUMN IF NOT EXISTS peso REAL DEFAULT 0")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tractomulas (
                    id SERIAL PRIMARY KEY,
                    placa TEXT UNIQUE NOT NULL,
                    consumo_km_galon REAL NOT NULL,
                    tipo TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS conductores (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT UNIQUE NOT NULL,
                    cedula TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS rutas (
                    id SERIAL PRIMARY KEY,
                    origen TEXT NOT NULL,
                    destino TEXT NOT NULL,
                    distancia_km REAL NOT NULL,
                    es_frontera INTEGER NOT NULL,
                    es_regional INTEGER NOT NULL,
                    es_aguachica INTEGER NOT NULL
                )
            ''')

            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS es_riohacha INTEGER DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_flypass REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_peajes REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_urea_acpm REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_hotel REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_comida REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_transporte REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_propina_comision REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_cargue_descargue REAL DEFAULT 0")
            cursor.execute("ALTER TABLE rutas ADD COLUMN IF NOT EXISTS default_otros REAL DEFAULT 0")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS liquidaciones_conductor (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    periodo_inicio DATE NOT NULL,
                    periodo_fin DATE NOT NULL,
                    viajes_incluidos TEXT,
                    cantidad_viajes INTEGER DEFAULT 0,
                    placas TEXT,
                    total_nomina REAL DEFAULT 0,
                    total_comisiones REAL DEFAULT 0,
                    total_anticipos REAL DEFAULT 0,
                    total_a_pagar REAL DEFAULT 0,
                    estado TEXT DEFAULT 'Pendiente',
                    fecha_pago DATE,
                    observaciones TEXT,
                    fecha_creacion TEXT
                )
            ''')
            cursor.execute("ALTER TABLE liquidaciones_conductor ADD COLUMN IF NOT EXISTS placas TEXT")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS cuentas_por_pagar_cobrar (
                    id SERIAL PRIMARY KEY,
                    tipo TEXT NOT NULL,
                    concepto TEXT NOT NULL,
                    tercero TEXT,
                    monto REAL NOT NULL,
                    fecha_vencimiento DATE NOT NULL,
                    estado TEXT DEFAULT 'Pendiente',
                    fecha_pago DATE,
                    viaje_id INTEGER,
                    observaciones TEXT,
                    fecha_creacion TEXT
                )
            ''')

            # ---------------- NUEVO v4.8: Días sin viaje (solo trazabilidad) ----------------
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS dias_sin_viaje (
                    id SERIAL PRIMARY KEY,
                    fecha DATE NOT NULL,
                    placa TEXT NOT NULL,
                    conductor TEXT,
                    motivo TEXT,
                    observaciones TEXT,
                    fecha_creacion TEXT
                )
            ''')

            conn.commit()
            _db_initialized = True
        except Exception as e:
            st.error(f"Error inicializando base de datos: {e}")
        finally:
            if conn is not None:
                self.release_connection(conn)

    def guardar_viaje(self, calculadora, fecha_viaje, observaciones="", cliente=""):
        conn = None
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            costos = calculadora.calcular_costos_totales()

            hora_colombia = datetime.now() - timedelta(hours=5)
            fecha_actual = hora_colombia.strftime('%Y-%m-%d %H:%M:%S')
            fecha_viaje_str = fecha_viaje.strftime('%Y-%m-%d') if fecha_viaje else fecha_actual[:10]

            datos_viaje = (
                fecha_actual,
                str(calculadora.tractomula.placa),
                str(calculadora.conductor.nombre),
                str(calculadora.ruta.origen),
                str(calculadora.ruta.destino),
                float(calculadora.distancia_efectiva),
                int(calculadora.dias_viaje),
                1 if calculadora.es_frontera else 0,
                1 if calculadora.hubo_parqueo else 0,
                costos['nomina_admin'],
                costos['nomina_conductor'],
                costos['comision_conductor'],
                costos['mantenimiento'],
                costos['seguros'],
                costos['tecnomecanica'],
                costos['llantas'],
                costos['aceite'],
                costos['combustible'],
                costos['galones_necesarios'],
                float(calculadora.flypass),
                float(calculadora.peajes),
                costos['cruce_frontera'],
                float(calculadora.hotel),
                float(calculadora.comida),
                costos['parqueo'],
                float(calculadora.cargue_descargue),
                float(calculadora.otros),
                costos['total_gastos'],
                costos['legalizacion'],
                costos['punto_equilibrio'],
                float(calculadora.valor_flete),
                costos['utilidad'],
                costos['rentabilidad'],
                float(calculadora.anticipo),
                costos['saldo'],
                1 if calculadora.hubo_anticipo_empresa else 0,
                costos['ant_empresa'],
                costos['saldo_empresa'],
                str(observaciones),
                float(calculadora.urea_acpm),
                float(calculadora.transporte),
                float(calculadora.propina_comision),
                fecha_viaje_str,
                str(cliente),
                int(calculadora.numero_viajes),
                float(calculadora.peso),
            )

            sql_insert = '''
                INSERT INTO viajes_v4 (
                    fecha_creacion, placa, conductor, origen, destino, distancia_km,
                    dias_viaje, es_frontera, hubo_parqueo, nomina_admin, nomina_conductor,
                    comision_conductor, mantenimiento, seguros, tecnomecanica, llantas,
                    aceite, combustible, galones_necesarios, flypass, peajes,
                    cruce_frontera, hotel, comida, parqueo, cargue_descargue, otros,
                    total_gastos, legalizacion, punto_equilibrio, valor_flete,
                    utilidad, rentabilidad, anticipo, saldo, hubo_anticipo_empresa,
                    ant_empresa, saldo_empresa, observaciones,
                    urea_acpm, transporte, propina_comision, fecha_viaje, cliente,
                    numero_viajes, peso
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s
                ) RETURNING id
            '''

            cursor.execute(sql_insert, datos_viaje)

            result = cursor.fetchone()
            if result:
                viaje_id = result[0]
            else:
                viaje_id = None
                st.warning("El viaje se guardó pero no se pudo recuperar el ID.")

            conn.commit()
            return viaje_id

        except Exception as e:
            st.error(f"❌ Error detallado al guardar: {e}")
            return None
        finally:
            if conn is not None:
                self.release_connection(conn)

    def obtener_todos_viajes(self):
        conn = self.get_connection()
        try:
            query = "SELECT * FROM viajes_v4 ORDER BY fecha_creacion DESC"
            df = pd.read_sql_query(query, conn)
            return df
        finally:
            self.release_connection(conn)

    def buscar_viajes(self, fecha_inicio=None, fecha_fin=None, placa=None, conductor=None, origen=None, destino=None, cliente=None):
        conn = self.get_connection()
        try:
            query = "SELECT * FROM viajes_v4 WHERE 1=1"
            params = []
            if fecha_inicio:
                query += " AND fecha_viaje >= %s"
                params.append(fecha_inicio)
            if fecha_fin:
                query += " AND fecha_viaje <= %s"
                params.append(fecha_fin)
            if placa:
                query += " AND placa = %s"
                params.append(placa)
            if conductor:
                query += " AND conductor ILIKE %s"
                params.append(f"%{conductor}%")
            if origen:
                query += " AND origen ILIKE %s"
                params.append(f"%{origen}%")
            if destino:
                query += " AND destino ILIKE %s"
                params.append(f"%{destino}%")
            if cliente:
                query += " AND cliente ILIKE %s"
                params.append(f"%{cliente}%")
            query += " ORDER BY fecha_creacion DESC"
            df = pd.read_sql_query(query, conn, params=params)
            return df
        finally:
            self.release_connection(conn)

    def obtener_viaje_por_id(self, viaje_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM viajes_v4 WHERE id = %s", (viaje_id,))
            viaje = cursor.fetchone()
            return viaje
        finally:
            self.release_connection(conn)

    def eliminar_viaje(self, viaje_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM viajes_v4 WHERE id = %s", (viaje_id,))
            conn.commit()
        finally:
            self.release_connection(conn)

    def actualizar_viaje(self, viaje_id, calculadora, fecha_viaje, observaciones="", cliente="", galones_reales=None):
        conn = None
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            costos = calculadora.calcular_costos_totales()
            fecha_viaje_str = fecha_viaje.strftime('%Y-%m-%d') if fecha_viaje else None
            galones_reales_val = float(galones_reales) if galones_reales else None

            cursor.execute('''
                UPDATE viajes_v4 SET
                    placa = %s, conductor = %s, origen = %s, destino = %s, distancia_km = %s,
                    dias_viaje = %s, es_frontera = %s, hubo_parqueo = %s,
                    nomina_admin = %s, nomina_conductor = %s, comision_conductor = %s,
                    mantenimiento = %s, seguros = %s, tecnomecanica = %s, llantas = %s,
                    aceite = %s, combustible = %s, galones_necesarios = %s,
                    flypass = %s, peajes = %s, cruce_frontera = %s, hotel = %s, comida = %s,
                    parqueo = %s, cargue_descargue = %s, otros = %s,
                    total_gastos = %s, legalizacion = %s, punto_equilibrio = %s, valor_flete = %s,
                    utilidad = %s, rentabilidad = %s, anticipo = %s, saldo = %s,
                    hubo_anticipo_empresa = %s, ant_empresa = %s, saldo_empresa = %s,
                    observaciones = %s, urea_acpm = %s, transporte = %s, propina_comision = %s,
                    fecha_viaje = %s, cliente = %s, galones_reales = %s, numero_viajes = %s,
                    peso = %s
                WHERE id = %s
            ''', (
                str(calculadora.tractomula.placa), str(calculadora.conductor.nombre),
                str(calculadora.ruta.origen), str(calculadora.ruta.destino),
                float(calculadora.distancia_efectiva), int(calculadora.dias_viaje),
                1 if calculadora.es_frontera else 0, 1 if calculadora.hubo_parqueo else 0,
                costos['nomina_admin'], costos['nomina_conductor'], costos['comision_conductor'],
                costos['mantenimiento'], costos['seguros'], costos['tecnomecanica'], costos['llantas'],
                costos['aceite'], costos['combustible'], costos['galones_necesarios'],
                float(calculadora.flypass), float(calculadora.peajes), costos['cruce_frontera'],
                float(calculadora.hotel), float(calculadora.comida), costos['parqueo'],
                float(calculadora.cargue_descargue), float(calculadora.otros),
                costos['total_gastos'], costos['legalizacion'], costos['punto_equilibrio'],
                float(calculadora.valor_flete), costos['utilidad'], costos['rentabilidad'],
                float(calculadora.anticipo), costos['saldo'],
                1 if calculadora.hubo_anticipo_empresa else 0, costos['ant_empresa'], costos['saldo_empresa'],
                str(observaciones), float(calculadora.urea_acpm), float(calculadora.transporte),
                float(calculadora.propina_comision), fecha_viaje_str, str(cliente), galones_reales_val,
                int(calculadora.numero_viajes), float(calculadora.peso), viaje_id
            ))
            conn.commit()
            return True
        except Exception as e:
            st.error(f"❌ Error al actualizar el viaje: {e}")
            return False
        finally:
            if conn is not None:
                self.release_connection(conn)

    def obtener_viajes_con_consumo(self, placa=None):
        conn = self.get_connection()
        try:
            query = """
                SELECT id, fecha_viaje, placa, conductor, origen, destino, distancia_km,
                       galones_necesarios, galones_reales, combustible
                FROM viajes_v4
                WHERE galones_reales IS NOT NULL AND galones_reales > 0
            """
            params = []
            if placa:
                query += " AND placa = %s"
                params.append(placa)
            query += " ORDER BY fecha_viaje DESC"
            df = pd.read_sql_query(query, conn, params=params)
        finally:
            self.release_connection(conn)

        if not df.empty:
            df['diferencia_galones'] = df['galones_reales'] - df['galones_necesarios']
            df['porcentaje_sobreconsumo'] = (
                df['diferencia_galones'] / df['galones_necesarios'] * 100
            ).where(df['galones_necesarios'] != 0, 0)
        return df

    def obtener_estadisticas(self):
        conn = self.get_connection()
        stats = {}
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(SUM(numero_viajes),0) FROM viajes_v4")
            stats['total_viajes'] = cursor.fetchone()[0]
            cursor.execute("SELECT SUM(distancia_km) FROM viajes_v4")
            stats['total_km'] = cursor.fetchone()[0] or 0
            cursor.execute("SELECT SUM(total_gastos) FROM viajes_v4")
            stats['total_gastos'] = cursor.fetchone()[0] or 0
            cursor.execute("SELECT placa, COALESCE(SUM(numero_viajes),0) as total FROM viajes_v4 GROUP BY placa ORDER BY total DESC")
            stats['viajes_por_placa'] = cursor.fetchall()
            cursor.execute("SELECT conductor, COALESCE(SUM(numero_viajes),0) as total FROM viajes_v4 GROUP BY conductor ORDER BY total DESC")
            stats['viajes_por_conductor'] = cursor.fetchall()
            cursor.execute("SELECT origen, destino, COALESCE(SUM(numero_viajes),0) as total FROM viajes_v4 GROUP BY origen, destino ORDER BY total DESC LIMIT 5")
            stats['rutas_frecuentes'] = cursor.fetchall()
        except Exception:
            stats = {'total_viajes': 0, 'total_km': 0, 'total_gastos': 0, 'viajes_por_placa': [], 'viajes_por_conductor': [], 'rutas_frecuentes': []}
        finally:
            self.release_connection(conn)
        return stats

    def obtener_dashboard_data(self):
        conn = self.get_connection()
        hoy = datetime.now()
        inicio_mes = hoy.replace(day=1).strftime('%Y-%m-%d')
        data = {}

        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT COALESCE(SUM(numero_viajes),0) as total_viajes, SUM(distancia_km) as total_km,
                       SUM(total_gastos) as total_gastos, SUM(valor_flete) as total_ingresos,
                       SUM(utilidad) as total_utilidad, AVG(utilidad) as utilidad_promedio
                FROM viajes_v4
                WHERE fecha_viaje >= %s
            """, (inicio_mes,))
            row = cursor.fetchone()
            data['mes_actual'] = {
                'total_viajes': row[0] or 0,
                'total_km': row[1] or 0,
                'total_gastos': row[2] or 0,
                'total_ingresos': row[3] or 0,
                'total_utilidad': row[4] or 0,
                'utilidad_promedio': row[5] or 0
            }

            cursor.execute("""
                SELECT placa, COALESCE(SUM(numero_viajes),0) as viajes, SUM(total_gastos) as gastos,
                       SUM(valor_flete) as ingresos, SUM(utilidad) as utilidad
                FROM viajes_v4
                WHERE fecha_viaje >= %s
                GROUP BY placa ORDER BY utilidad DESC
            """, (inicio_mes,))
            data['por_tractomula'] = cursor.fetchall()

            cursor.execute("""
                SELECT conductor, COALESCE(SUM(numero_viajes),0) as viajes, SUM(utilidad) as utilidad,
                       AVG(utilidad) as utilidad_promedio
                FROM viajes_v4
                WHERE fecha_viaje >= %s
                GROUP BY conductor ORDER BY utilidad DESC
            """, (inicio_mes,))
            data['por_conductor'] = cursor.fetchall()

            cursor.execute("""
                SELECT origen, destino, COALESCE(SUM(numero_viajes),0) as viajes, AVG(utilidad) as utilidad_promedio,
                       SUM(utilidad) as utilidad_total
                FROM viajes_v4
                WHERE fecha_viaje >= %s
                GROUP BY origen, destino ORDER BY utilidad_total DESC LIMIT 5
            """, (inicio_mes,))
            data['rutas_rentables'] = cursor.fetchall()

            cursor.execute("""
                SELECT to_char(fecha_viaje, 'YYYY-MM') as mes,
                       COALESCE(SUM(numero_viajes),0) as viajes, SUM(total_gastos) as gastos,
                       SUM(valor_flete) as ingresos, SUM(utilidad) as utilidad
                FROM viajes_v4
                WHERE fecha_viaje >= CURRENT_DATE - INTERVAL '6 months'
                GROUP BY mes ORDER BY mes
            """)
            data['evolucion_6_meses'] = cursor.fetchall()

            cursor.execute("""
                SELECT fecha_creacion, placa, origen, destino, total_gastos, valor_flete, utilidad
                FROM viajes_v4 WHERE utilidad < 0 ORDER BY fecha_creacion DESC LIMIT 10
            """)
            data['viajes_no_rentables'] = cursor.fetchall()

            cursor.execute("""
                SELECT SUM(valor_flete) as ut_bruta, SUM(utilidad) as ut_neta
                FROM viajes_v4 WHERE fecha_viaje >= %s
            """, (inicio_mes,))
            row_ut = cursor.fetchone()
            ut_bruta = row_ut[0] or 0
            ut_neta = row_ut[1] or 0
            porcentaje_ut = (ut_neta / ut_bruta * 100) if ut_bruta > 0 else 0
            data['ut_bruta'] = ut_bruta
            data['ut_neta'] = ut_neta
            data['porcentaje_ut'] = porcentaje_ut

        except Exception:
            data = {k: 0 for k in ['ut_bruta', 'ut_neta', 'porcentaje_ut']}
            data['mes_actual'] = {k: 0 for k in ['total_viajes', 'total_km', 'total_gastos', 'total_ingresos', 'total_utilidad', 'utilidad_promedio']}
            data['por_tractomula'] = []
            data['por_conductor'] = []
            data['rutas_rentables'] = []
            data['evolucion_6_meses'] = []
            data['viajes_no_rentables'] = []
        finally:
            self.release_connection(conn)

        return data

    def obtener_totales_por_placa(self, fecha_inicio=None, fecha_fin=None):
        conn = self.get_connection()
        try:
            query = """
                SELECT placa, SUM(valor_flete) as total_cxc, SUM(nomina_admin) as total_admin,
                       SUM(nomina_conductor) as total_parafiscales, SUM(comision_conductor) as total_comision,
                       SUM(mantenimiento) as total_mantenimiento, SUM(seguros) as total_seguros,
                       SUM(tecnomecanica) as total_tecnomecanica, SUM(llantas) as total_llantas,
                       SUM(aceite) as total_aceite, SUM(combustible) as total_combustible,
                       SUM(flypass) as total_flypass, SUM(peajes) as total_peajes,
                       SUM(urea_acpm) as total_urea_acpm,
                       SUM(cruce_frontera) as total_cruce_frontera, SUM(hotel) as total_hotel,
                       SUM(comida) as total_comida, SUM(transporte) as total_transporte,
                       SUM(parqueo) as total_parqueo, SUM(propina_comision) as total_propina_comision,
                       SUM(cargue_descargue) as total_cargue_descargue, SUM(otros) as total_otros,
                       SUM(legalizacion) as total_legalizacion, SUM(anticipo) as total_anticipos,
                       SUM(saldo) as total_saldo, SUM(ant_empresa) as total_ant_empresa,
                       SUM(saldo_empresa) as total_saldo_empresa
                FROM viajes_v4 WHERE 1=1
            """
            params = []
            if fecha_inicio:
                query += " AND fecha_viaje >= %s"
                params.append(fecha_inicio)
            if fecha_fin:
                query += " AND fecha_viaje <= %s"
                params.append(fecha_fin)
            query += " GROUP BY placa ORDER BY placa"

            try:
                df = pd.read_sql_query(query, conn, params=params)
                if not df.empty:
                    df['total_gastos'] = (
                        df['total_admin'] + df['total_parafiscales'] + df['total_comision'] +
                        df['total_mantenimiento'] + df['total_seguros'] + df['total_tecnomecanica'] +
                        df['total_llantas'] + df['total_aceite'] + df['total_combustible'] +
                        df['total_flypass'] + df['total_peajes'] + df['total_urea_acpm'] +
                        df['total_cruce_frontera'] + df['total_hotel'] + df['total_comida'] +
                        df['total_transporte'] + df['total_parqueo'] + df['total_propina_comision'] +
                        df['total_cargue_descargue'] + df['total_otros']
                    )
                    df['total_punto_equilibrio'] = df['total_cxc'] * 0.40
                    df['total_ut'] = df['total_cxc'] - df['total_gastos']
                    df['total_rentabilidad'] = (df['total_ut'] / df['total_cxc'] * 100).where(df['total_cxc'] != 0, 0)
                    df['total_saldo'] = df['total_anticipos'] - df['total_legalizacion']
            except Exception:
                df = pd.DataFrame()

            return df
        finally:
            self.release_connection(conn)

    # Métodos para tractomulas
    def guardar_tractomula(self, tractomula):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO tractomulas (placa, consumo_km_galon, tipo)
                VALUES (%s, %s, %s)
            ''', (tractomula.placa, tractomula.consumo_km_galon, tractomula.tipo))
            conn.commit()
            return True
        except Exception:
            return False
        finally:
            self.release_connection(conn)

    def obtener_tractomulas(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM tractomulas ORDER BY placa")
            tractomulas = []
            for row in cursor.fetchall():
                tractomulas.append(Tractomula(
                    placa=row[1],
                    consumo_km_galon=row[2],
                    tipo=row[3]
                ))
            return tractomulas
        finally:
            self.release_connection(conn)

    def eliminar_tractomula(self, placa):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM tractomulas WHERE placa = %s", (placa,))
            conn.commit()
        finally:
            self.release_connection(conn)

    # Métodos para conductores
    def guardar_conductor(self, conductor):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO conductores (nombre, cedula)
                VALUES (%s, %s)
            ''', (conductor.nombre, conductor.cedula))
            conn.commit()
            return True
        except Exception:
            return False
        finally:
            self.release_connection(conn)

    def obtener_conductores(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM conductores ORDER BY nombre")
            conductores = []
            for row in cursor.fetchall():
                conductores.append(Conductor(nombre=row[1], cedula=row[2]))
            return conductores
        finally:
            self.release_connection(conn)

    def eliminar_conductor(self, nombre):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM conductores WHERE nombre = %s", (nombre,))
            conn.commit()
        finally:
            self.release_connection(conn)

    # ---------------- Métodos para rutas ----------------
    def guardar_ruta(self, ruta):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO rutas (
                    origen, destino, distancia_km, es_frontera, es_regional, es_aguachica,
                    es_riohacha, default_flypass, default_peajes, default_urea_acpm,
                    default_hotel, default_comida, default_transporte,
                    default_propina_comision, default_cargue_descargue, default_otros
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                ruta.origen,
                ruta.destino,
                ruta.distancia_km,
                1 if ruta.es_frontera else 0,
                1 if ruta.es_regional else 0,
                1 if ruta.es_aguachica else 0,
                1 if ruta.es_riohacha else 0,
                ruta.default_flypass,
                ruta.default_peajes,
                ruta.default_urea_acpm,
                ruta.default_hotel,
                ruta.default_comida,
                ruta.default_transporte,
                ruta.default_propina_comision,
                ruta.default_cargue_descargue,
                ruta.default_otros,
            ))
            result = cursor.fetchone()
            ruta_id = result[0] if result else None
            conn.commit()
            return ruta_id
        finally:
            self.release_connection(conn)

    def obtener_rutas(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, origen, destino, distancia_km, es_frontera, es_regional, es_aguachica,
                       es_riohacha, default_flypass, default_peajes, default_urea_acpm,
                       default_hotel, default_comida, default_transporte,
                       default_propina_comision, default_cargue_descargue, default_otros
                FROM rutas
                ORDER BY origen, destino
            """)

            rutas = []
            for row in cursor.fetchall():
                rutas.append(Ruta(
                    origen=row[1],
                    destino=row[2],
                    distancia_km=row[3],
                    es_frontera=bool(row[4]),
                    es_regional=bool(row[5]),
                    es_aguachica=bool(row[6]),
                    es_riohacha=bool(row[7]),
                    default_flypass=row[8] or 0.0,
                    default_peajes=row[9] or 0.0,
                    default_urea_acpm=row[10] or 0.0,
                    default_hotel=row[11] or 0.0,
                    default_comida=row[12] or 0.0,
                    default_transporte=row[13] or 0.0,
                    default_propina_comision=row[14] or 0.0,
                    default_cargue_descargue=row[15] or 0.0,
                    default_otros=row[16] or 0.0,
                ))
            return rutas
        finally:
            self.release_connection(conn)

    def obtener_rutas_con_id(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, origen, destino, distancia_km, es_frontera, es_regional, es_aguachica, es_riohacha
                FROM rutas
                ORDER BY origen, destino
            """)
            rutas_con_id = cursor.fetchall()
            return rutas_con_id
        finally:
            self.release_connection(conn)

    def eliminar_ruta(self, ruta_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM rutas WHERE id = %s", (ruta_id,))
            conn.commit()
        finally:
            self.release_connection(conn)

    # ---------------- Métodos para Liquidaciones de Conductores ----------------
    def obtener_viajes_para_liquidar(self, conductor, periodo_inicio, periodo_fin):
        conn = self.get_connection()
        try:
            query = """
                SELECT id, fecha_viaje, placa, origen, destino, comision_conductor, numero_viajes
                FROM viajes_v4
                WHERE conductor = %s AND fecha_viaje >= %s AND fecha_viaje <= %s
                ORDER BY fecha_viaje
            """
            df = pd.read_sql_query(query, conn, params=[conductor, periodo_inicio, periodo_fin])
            return df
        finally:
            self.release_connection(conn)

    def guardar_liquidacion(self, conductor, periodo_inicio, periodo_fin, df_viajes, observaciones=""):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()

            total_comisiones = float(df_viajes['comision_conductor'].sum()) if not df_viajes.empty else 0.0
            total_a_pagar = total_comisiones
            viajes_incluidos = ",".join(str(i) for i in df_viajes['id'].tolist()) if not df_viajes.empty else ""
            cantidad_viajes = int(df_viajes['numero_viajes'].fillna(1).sum()) if not df_viajes.empty else 0
            placas = ", ".join(sorted(df_viajes['placa'].unique())) if not df_viajes.empty else ""

            hora_colombia = datetime.now() - timedelta(hours=5)
            fecha_creacion = hora_colombia.strftime('%Y-%m-%d %H:%M:%S')

            cursor.execute('''
                INSERT INTO liquidaciones_conductor (
                    conductor, periodo_inicio, periodo_fin, viajes_incluidos, cantidad_viajes, placas,
                    total_comisiones, total_a_pagar,
                    estado, observaciones, fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s)
                RETURNING id
            ''', (conductor, periodo_inicio, periodo_fin, viajes_incluidos, cantidad_viajes, placas,
                  total_comisiones, total_a_pagar,
                  observaciones, fecha_creacion))
            result = cursor.fetchone()
            liquidacion_id = result[0] if result else None
            conn.commit()
            return liquidacion_id, cantidad_viajes, placas, total_comisiones, total_a_pagar
        finally:
            self.release_connection(conn)

    def obtener_liquidaciones(self, conductor=None, estado=None):
        conn = self.get_connection()
        try:
            query = "SELECT * FROM liquidaciones_conductor WHERE 1=1"
            params = []
            if conductor:
                query += " AND conductor = %s"
                params.append(conductor)
            if estado:
                query += " AND estado = %s"
                params.append(estado)
            query += " ORDER BY periodo_inicio DESC"
            df = pd.read_sql_query(query, conn, params=params)
            return df
        finally:
            self.release_connection(conn)

    def marcar_liquidacion_pagada(self, liquidacion_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            hoy = (datetime.now() - timedelta(hours=5)).strftime('%Y-%m-%d')
            cursor.execute("UPDATE liquidaciones_conductor SET estado = 'Pagada', fecha_pago = %s WHERE id = %s",
                           (hoy, liquidacion_id))
            conn.commit()
        finally:
            self.release_connection(conn)

    def eliminar_liquidacion(self, liquidacion_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM liquidaciones_conductor WHERE id = %s", (liquidacion_id,))
            conn.commit()
        finally:
            self.release_connection(conn)

    # ---------------- Métodos para Cuentas por Pagar/Cobrar ----------------
    def guardar_cuenta(self, tipo, concepto, tercero, monto, fecha_vencimiento, observaciones="", viaje_id=None):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            hora_colombia = datetime.now() - timedelta(hours=5)
            fecha_creacion = hora_colombia.strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                INSERT INTO cuentas_por_pagar_cobrar (
                    tipo, concepto, tercero, monto, fecha_vencimiento, estado, observaciones, viaje_id, fecha_creacion
                ) VALUES (%s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s)
                RETURNING id
            ''', (tipo, concepto, tercero, monto, fecha_vencimiento, observaciones, viaje_id, fecha_creacion))
            result = cursor.fetchone()
            cuenta_id = result[0] if result else None
            conn.commit()
            return cuenta_id
        finally:
            self.release_connection(conn)

    def obtener_cuentas(self, tipo=None, estado=None):
        conn = self.get_connection()
        try:
            query = "SELECT * FROM cuentas_por_pagar_cobrar WHERE 1=1"
            params = []
            if tipo:
                query += " AND tipo = %s"
                params.append(tipo)
            if estado:
                query += " AND estado = %s"
                params.append(estado)
            query += " ORDER BY fecha_vencimiento ASC"
            df = pd.read_sql_query(query, conn, params=params)
            return df
        finally:
            self.release_connection(conn)

    def marcar_cuenta_pagada(self, cuenta_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            hoy = (datetime.now() - timedelta(hours=5)).strftime('%Y-%m-%d')
            cursor.execute("UPDATE cuentas_por_pagar_cobrar SET estado = 'Pagado', fecha_pago = %s WHERE id = %s",
                           (hoy, cuenta_id))
            conn.commit()
        finally:
            self.release_connection(conn)

    def eliminar_cuenta(self, cuenta_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM cuentas_por_pagar_cobrar WHERE id = %s", (cuenta_id,))
            conn.commit()
        finally:
            self.release_connection(conn)

    # ---------------- NUEVO v4.8: Métodos para Días Sin Viaje ----------------
    def guardar_dia_sin_viaje(self, fecha, placa, conductor="", motivo="", observaciones=""):
        """Registra un día en el que la tractomula NO hizo viaje. Solo para trazabilidad,
        no calcula ni guarda ningún costo, gasto, flete o utilidad."""
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            hora_colombia = datetime.now() - timedelta(hours=5)
            fecha_creacion = hora_colombia.strftime('%Y-%m-%d %H:%M:%S')
            fecha_str = fecha.strftime('%Y-%m-%d') if hasattr(fecha, 'strftime') else fecha
            cursor.execute('''
                INSERT INTO dias_sin_viaje (fecha, placa, conductor, motivo, observaciones, fecha_creacion)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (fecha_str, placa, conductor, motivo, observaciones, fecha_creacion))
            result = cursor.fetchone()
            dia_id = result[0] if result else None
            conn.commit()
            return dia_id
        except Exception as e:
            st.error(f"❌ Error al guardar el día sin viaje: {e}")
            return None
        finally:
            self.release_connection(conn)

    def obtener_dias_sin_viaje(self, placa=None, fecha_inicio=None, fecha_fin=None):
        conn = self.get_connection()
        try:
            query = "SELECT * FROM dias_sin_viaje WHERE 1=1"
            params = []
            if placa:
                query += " AND placa = %s"
                params.append(placa)
            if fecha_inicio:
                query += " AND fecha >= %s"
                params.append(fecha_inicio)
            if fecha_fin:
                query += " AND fecha <= %s"
                params.append(fecha_fin)
            query += " ORDER BY fecha DESC"
            df = pd.read_sql_query(query, conn, params=params)
            return df
        finally:
            self.release_connection(conn)

    def eliminar_dia_sin_viaje(self, dia_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM dias_sin_viaje WHERE id = %s", (dia_id,))
            conn.commit()
        finally:
            self.release_connection(conn)


# ==================== CLASES DE DATOS ====================
@dataclass
class Tractomula:
    placa: str
    consumo_km_galon: float
    tipo: str


@dataclass
class Conductor:
    nombre: str
    cedula: str


@dataclass
class Ruta:
    origen: str
    destino: str
    distancia_km: float
    es_frontera: bool
    es_regional: bool = False
    es_aguachica: bool = False
    es_riohacha: bool = False
    default_flypass: float = 0.0
    default_peajes: float = 0.0
    default_urea_acpm: float = 0.0
    default_hotel: float = 0.0
    default_comida: float = 0.0
    default_transporte: float = 0.0
    default_propina_comision: float = 0.0
    default_cargue_descargue: float = 0.0
    default_otros: float = 0.0

    @property
    def es_urbana(self) -> bool:
        return not (self.es_frontera or self.es_regional or self.es_aguachica or self.es_riohacha)


# ==================== DATOS COLOMBIANOS ====================
class DatosColombia:
    PRECIO_DIESEL = 10800
    NOMINA_ADMIN_BASE = 1300000
    NOMINA_ADMIN_DIVISOR = 14
    NOMINA_CONDUCTOR_DIA = 20000

    COMISION_URBANO_DIA = 120000
    COMISION_FRONTERA = 565000
    COMISION_REGIONAL = 200000
    COMISION_AGUACHICA = 360000
    COMISION_RIOACHA = 350000

    MANTENIMIENTO_MENSUAL = 1500000
    SEGURO_1 = 1400000
    SEGURO_2 = 6000000
    SEGURO_3 = 16000000
    TECNOMECANICA_ANUAL = 460000
    LLANTAS_COSTO = 1300000
    LLANTAS_CANTIDAD = 22
    LLANTAS_KM = 80000
    ACEITE_COSTO = 2500000
    ACEITE_KM = 15000
    CRUCE_FRONTERA = 560000
    PARQUEO_DIA = 15000
    MARGEN_ANT_EMPRESA = 0.90
    PUNTO_EQUILIBRIO_PORCENTAJE = 0.40

    AGOFER_VALOR_POR_KG = 27500
    AGOFER_CARGUE_DESCARGUE = 30000


# ==================== ASIGNACION DE CONDUCTORES ====================
PLACA_CONDUCTOR = {
    "NOX459": "GONZALO PINTO",
    "NOX460": "JOSE ORTEGA PEREZ",
    "NOX461": "ALVARO TAFUR",
    "SON047": "ISAIAS VESGA",
    "SON048": "FLAVIO ROSENDO MALTE TUTALCHA",
    "SOP148": "SLITH JOSE ORTEGA PACHECO",
    "SOP149": "ABRAHAM SEGUNDO ALVAREZ VALLE",
    "SOP150": "RAMON TAFUR HERNANDEZ",
    "SRO661": "",
    "SRO672": "PEDRO VILLAMIL",
    "TMW882": "JESUS DAVID MONTE MOSQUERA",
    "TRL282": "CHRISTIAN MARTINEZ NAVARRO",
    "TRL298": "YEIMI DUQUE ZULUAGA",
    "UYQ308": "JULIAN CALETH CORONADO",
    "UYV084": "CARLOS TAFUR",
    "UYY788": "EDUARDO RAFAEL OLIVARES ALCAZAR",
    "WLP822": "REIMUR VILLAMIL",
}


# ==================== CALCULADORA DE COSTOS ====================
class CalculadoraCostos:
    """Calcula todos los costos del viaje con fórmulas ACTUALIZADAS v4.3"""

    def __init__(self, tractomula: Tractomula, conductor: Conductor, ruta: Ruta,
                 dias_viaje: int, numero_viajes: int, es_frontera: bool, hubo_parqueo: bool,
                 flypass: float, peajes: float, urea_acpm: float, hotel: float,
                 comida: float, transporte: float, propina_comision: float,
                 cargue_descargue: float, otros: float, valor_flete: float,
                 anticipo: float, hubo_anticipo_empresa: bool, datos: DatosColombia,
                 peso: float = 0.0, cliente: str = ""):
        self.tractomula = tractomula
        self.conductor = conductor
        self.ruta = ruta
        self.dias_viaje = dias_viaje
        self.numero_viajes = numero_viajes
        self.es_frontera = es_frontera
        self.hubo_parqueo = hubo_parqueo
        self.flypass = flypass
        self.peajes = peajes
        self.urea_acpm = urea_acpm
        self.hotel = hotel
        self.comida = comida
        self.transporte = transporte
        self.propina_comision = propina_comision
        self.cargue_descargue = cargue_descargue
        self.otros = otros
        self.valor_flete = valor_flete
        self.anticipo = anticipo
        self.hubo_anticipo_empresa = hubo_anticipo_empresa
        self.datos = datos
        self.peso = peso
        self.cliente = cliente

    def aplica_formula_agofer(self) -> bool:
        return self.ruta.es_urbana and not self.es_frontera and es_cliente_agofer(self.cliente)

    @property
    def distancia_efectiva(self) -> float:
        if self.aplica_formula_agofer():
            return self.ruta.distancia_km * self.numero_viajes
        return self.ruta.distancia_km

    def calcular_flete_sugerido_agofer(self) -> float:
        return self.peso * self.datos.AGOFER_VALOR_POR_KG * self.numero_viajes

    def calcular_cargue_descargue_sugerido_agofer(self) -> float:
        return self.datos.AGOFER_CARGUE_DESCARGUE * self.numero_viajes

    def calcular_nomina_admin(self) -> float:
        return (self.datos.NOMINA_ADMIN_BASE / self.datos.NOMINA_ADMIN_DIVISOR) * self.dias_viaje

    def calcular_nomina_conductor(self) -> float:
        return self.datos.NOMINA_CONDUCTOR_DIA * self.dias_viaje

    def calcular_comision_conductor(self) -> float:
        if self.ruta.es_aguachica:
            return self.datos.COMISION_AGUACHICA
        elif self.ruta.es_riohacha:
            return self.datos.COMISION_RIOACHA
        elif self.ruta.es_regional:
            return self.datos.COMISION_REGIONAL
        elif self.es_frontera:
            return self.datos.COMISION_FRONTERA
        else:
            return self.datos.COMISION_URBANO_DIA * self.numero_viajes

    def calcular_mantenimiento(self) -> float:
        return (self.datos.MANTENIMIENTO_MENSUAL / 30) * self.dias_viaje

    def calcular_seguros(self) -> float:
        seguro_diario = (
            (self.datos.SEGURO_1 / 365) +
            (self.datos.SEGURO_2 / 365) +
            (self.datos.SEGURO_3 / 14 / 365)
        )
        return seguro_diario * self.dias_viaje

    def calcular_tecnomecanica(self) -> float:
        return (self.datos.TECNOMECANICA_ANUAL / 365) * self.dias_viaje

    def calcular_llantas(self) -> float:
        costo_por_km = (self.datos.LLANTAS_COSTO * self.datos.LLANTAS_CANTIDAD) / self.datos.LLANTAS_KM
        return costo_por_km * self.distancia_efectiva

    def calcular_aceite(self) -> float:
        costo_por_km = self.datos.ACEITE_COSTO / self.datos.ACEITE_KM
        return costo_por_km * self.distancia_efectiva

    def calcular_galones_necesarios(self) -> float:
        if self.tractomula.consumo_km_galon <= 0:
            return 0.0
        return self.distancia_efectiva / self.tractomula.consumo_km_galon

    def calcular_combustible(self) -> float:
        galones = self.calcular_galones_necesarios()
        return galones * self.datos.PRECIO_DIESEL

    def calcular_cruce_frontera(self) -> float:
        return self.datos.CRUCE_FRONTERA if self.es_frontera else 0

    def calcular_parqueo(self) -> float:
        return self.datos.PARQUEO_DIA * self.dias_viaje if self.hubo_parqueo else 0

    def calcular_legalizacion(self) -> float:
        return (self.peajes + self.urea_acpm + self.calcular_cruce_frontera() + self.hotel +
                self.comida + self.transporte + self.calcular_parqueo() +
                self.propina_comision + self.cargue_descargue + self.otros)

    def calcular_saldo(self) -> float:
        legalizacion = self.calcular_legalizacion()
        return self.anticipo - legalizacion

    def calcular_ant_empresa(self) -> float:
        if self.hubo_anticipo_empresa:
            return self.valor_flete * self.datos.MARGEN_ANT_EMPRESA
        else:
            return 0.0

    def calcular_costos_totales(self) -> Dict[str, float]:
        nomina_admin = self.calcular_nomina_admin()
        nomina_conductor = self.calcular_nomina_conductor()
        comision_conductor = self.calcular_comision_conductor()
        mantenimiento = self.calcular_mantenimiento()
        seguros = self.calcular_seguros()
        tecnomecanica = self.calcular_tecnomecanica()
        llantas = self.calcular_llantas()
        aceite = self.calcular_aceite()
        galones_necesarios = self.calcular_galones_necesarios()
        combustible = self.calcular_combustible()
        cruce_frontera = self.calcular_cruce_frontera()
        parqueo = self.calcular_parqueo()

        total_gastos = (
            nomina_admin + nomina_conductor + comision_conductor + mantenimiento +
            seguros + tecnomecanica + llantas + aceite + combustible +
            self.flypass + self.peajes + self.urea_acpm + cruce_frontera + self.hotel +
            self.comida + self.transporte + parqueo + self.propina_comision +
            self.cargue_descargue + self.otros
        )

        legalizacion = self.calcular_legalizacion()
        saldo = self.calcular_saldo()

        punto_equilibrio = self.valor_flete * self.datos.PUNTO_EQUILIBRIO_PORCENTAJE

        utilidad = self.valor_flete - total_gastos
        rentabilidad = (utilidad / self.valor_flete * 100) if self.valor_flete > 0 else 0
        ant_empresa = self.calcular_ant_empresa()
        saldo_empresa = self.valor_flete - ant_empresa

        return {
            'nomina_admin': round(nomina_admin, 2),
            'nomina_conductor': round(nomina_conductor, 2),
            'comision_conductor': round(comision_conductor, 2),
            'mantenimiento': round(mantenimiento, 2),
            'seguros': round(seguros, 2),
            'tecnomecanica': round(tecnomecanica, 2),
            'llantas': round(llantas, 2),
            'aceite': round(aceite, 2),
            'combustible': round(combustible, 2),
            'galones_necesarios': round(galones_necesarios, 2),
            'cruce_frontera': round(cruce_frontera, 2),
            'parqueo': round(parqueo, 2),
            'total_gastos': round(total_gastos, 2),
            'legalizacion': round(legalizacion, 2),
            'saldo': round(saldo, 2),
            'punto_equilibrio': round(punto_equilibrio, 2),
            'utilidad': round(utilidad, 2),
            'rentabilidad': round(rentabilidad, 2),
            'ant_empresa': round(ant_empresa, 2),
            'saldo_empresa': round(saldo_empresa, 2),
        }


# ==================== GENERADOR DE REPORTES ====================
class GeneradorReportes:
    """Genera reportes detallados de costos"""

    @staticmethod
    def generar_reporte_texto(calculadora: CalculadoraCostos) -> str:
        costos = calculadora.calcular_costos_totales()

        reporte = f"""
{'='*70}
          REPORTE DE COSTOS - TRANSPORTE DE CARGA
{'='*70}
INFORMACIÓN DEL VIAJE
{'-'*70}
Ruta: {calculadora.ruta.origen} → {calculadora.ruta.destino}
Distancia de la ruta: {formatear_numero(calculadora.ruta.distancia_km)} km
Distancia efectiva del día (ruta x N° viajes si aplica): {formatear_numero(calculadora.distancia_efectiva)} km
Días del viaje: {calculadora.dias_viaje}
Número de viajes: {calculadora.numero_viajes}
Peso transportado: {formatear_numero(calculadora.peso)} kg
Cliente: {calculadora.cliente or '-'}
Galones necesarios: {formatear_decimal(costos['galones_necesarios'])} gal
Es frontera: {'Sí' if calculadora.es_frontera else 'No'}
Es regional: {'Sí' if calculadora.ruta.es_regional else 'No'}
Es Aguachica: {'Sí' if calculadora.ruta.es_aguachica else 'No'}
Es Riohacha: {'Sí' if calculadora.ruta.es_riohacha else 'No'}
Hubo parqueo: {'Sí' if calculadora.hubo_parqueo else 'No'}

VEHÍCULO
{'-'*70}
Placa: {calculadora.tractomula.placa}
Tipo: {calculadora.tractomula.tipo}
Consumo usado en este viaje: {calculadora.obtener_consumo_km_galon()} km/galón

CONDUCTOR
{'-'*70}
Nombre: {calculadora.conductor.nombre}
Cédula: {calculadora.conductor.cedula}

DESGLOSE DE COSTOS
{'='*70}

1. Nómina Admin:          ${formatear_numero(costos['nomina_admin']):>18} COP
2. Nómina Conductor:      ${formatear_numero(costos['nomina_conductor']):>18} COP
3. Comisión Conductor:    ${formatear_numero(costos['comision_conductor']):>18} COP
4. Mantenimiento:         ${formatear_numero(costos['mantenimiento']):>18} COP
5. Seguros:               ${formatear_numero(costos['seguros']):>18} COP
6. Tecnomecánica:         ${formatear_numero(costos['tecnomecanica']):>18} COP
7. Llantas:               ${formatear_numero(costos['llantas']):>18} COP
8. Aceite:                ${formatear_numero(costos['aceite']):>18} COP
9. Combustible:           ${formatear_numero(costos['combustible']):>18} COP
10. Flypass:              ${formatear_numero(calculadora.flypass):>18} COP
11. Peajes:               ${formatear_numero(calculadora.peajes):>18} COP
12. Urea y/o ACPM:        ${formatear_numero(calculadora.urea_acpm):>18} COP
13. Cruce Frontera:       ${formatear_numero(costos['cruce_frontera']):>18} COP
14. Hotel:                ${formatear_numero(calculadora.hotel):>18} COP
15. Comida:                ${formatear_numero(calculadora.comida):>18} COP
16. Transporte:           ${formatear_numero(calculadora.transporte):>18} COP
17. Parqueo:              ${formatear_numero(costos['parqueo']):>18} COP
18. Propina/Comisión:     ${formatear_numero(calculadora.propina_comision):>18} COP
19. Cargue/Descargue:     ${formatear_numero(calculadora.cargue_descargue):>18} COP
20. Otros (engrase, etc): ${formatear_numero(calculadora.otros):>18} COP
{'='*70}

RESULTADOS
{'='*70}
TOTAL GASTOS:             ${formatear_numero(costos['total_gastos']):>18} COP
LEGALIZACIÓN:             ${formatear_numero(costos['legalizacion']):>18} COP
ANTICIPO:                 ${formatear_numero(calculadora.anticipo):>18} COP
SALDO:                    ${formatear_numero(costos['saldo']):>18} COP
PUNTO DE EQUILIBRIO:      ${formatear_numero(costos['punto_equilibrio']):>18} COP
VALOR DEL FLETE:          ${formatear_numero(calculadora.valor_flete):>18} COP
UTILIDAD (UT):            ${formatear_numero(costos['utilidad']):>18} COP
RENTABILIDAD:             {costos['rentabilidad']:>18.1f} %
ANT. EMPRESA (90%):       ${formatear_numero(costos['ant_empresa']):>18} COP
SALDO EMPRESA:            ${formatear_numero(costos['saldo_empresa']):>18} COP

Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*70}
        """
        return reporte

    @staticmethod
    def generar_excel(calculadoras: List[CalculadoraCostos]) -> io.BytesIO:
        """Genera un archivo Excel en memoria para descarga"""
        output = io.BytesIO()
        wb = Workbook()

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"

        ws_resumen.merge_cells('A1:O1')
        cell = ws_resumen['A1']
        cell.value = "REPORTE DE COSTOS - TRANSPORTE DE CARGA COLOMBIA"
        cell.font = Font(size=14, bold=True, color="1F4E78")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_resumen.merge_cells('A2:O2')
        ws_resumen['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_resumen['A2'].alignment = Alignment(horizontal='center')

        row = 4
        headers = ['Ruta', 'Placa', 'Conductor', 'Peso (kg)', 'Distancia (km)', 'Días', 'N° Viajes', 'Galones',
                   'Combustible', 'Total Gastos', 'Anticipo', 'Saldo', 'Valor Flete',
                   'Utilidad', 'Rentabilidad %']

        for col, header in enumerate(headers, start=1):
            cell = ws_resumen.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row = 5
        for calc in calculadoras:
            costos = calc.calcular_costos_totales()
            ruta_str = f"{calc.ruta.origen} → {calc.ruta.destino}"

            datos = [
                ruta_str,
                calc.tractomula.placa,
                calc.conductor.nombre,
                calc.peso,
                calc.distancia_efectiva,
                calc.dias_viaje,
                calc.numero_viajes,
                costos['galones_necesarios'],
                costos['combustible'],
                costos['total_gastos'],
                calc.anticipo,
                costos['saldo'],
                calc.valor_flete,
                costos['utilidad'],
                costos['rentabilidad']
            ]

            for col, valor in enumerate(datos, start=1):
                cell = ws_resumen.cell(row=row, column=col)
                cell.value = valor
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right')
                if col == 4:
                    cell.number_format = '#,##0'
                elif col == 5:
                    cell.number_format = '#,##0'
                elif col == 6:
                    cell.number_format = '#,##0'
                elif col == 7:
                    cell.number_format = '#,##0'
                elif col == 8:
                    cell.number_format = '#,##0.00'
                elif col >= 9 and col <= 14:
                    cell.number_format = '$#,##0'
                elif col == 15:
                    cell.number_format = '#,##0.0"%"'
                else:
                    cell.number_format = '#,##0'

            row += 1

        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 12
        ws_resumen.column_dimensions['C'].width = 28
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws_resumen.column_dimensions[col].width = 16

        for idx, calc in enumerate(calculadoras, start=1):
            costos = calc.calcular_costos_totales()
            ws = wb.create_sheet(title=f"Ruta {idx}")

            ws.merge_cells('A1:D1')
            ws['A1'] = f"{calc.ruta.origen} → {calc.ruta.destino}"
            ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
            ws['A1'].alignment = Alignment(horizontal='center')

            ws['A2'] = f"Días: {calc.dias_viaje}   |   Número de viajes: {calc.numero_viajes}   |   Peso: {formatear_numero(calc.peso)} kg   |   Distancia efectiva: {formatear_numero(calc.distancia_efectiva)} km"
            ws['A2'].font = Font(italic=True, color="555555")

            row = 3

            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DESGLOSE DE COSTOS"
            cell.font = subheader_font
            cell.fill = subheader_fill
            row += 1

            ws[f'A{row}'] = "Concepto"
            ws[f'B{row}'] = "Monto (COP)"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            detalles_costos = [
                ('1. Nómina Admin', costos['nomina_admin']),
                ('2. Nómina Conductor', costos['nomina_conductor']),
                ('3. Comisión Conductor', costos['comision_conductor']),
                ('4. Mantenimiento', costos['mantenimiento']),
                ('5. Seguros', costos['seguros']),
                ('6. Tecnomecánica', costos['tecnomecanica']),
                ('7. Llantas', costos['llantas']),
                ('8. Aceite', costos['aceite']),
                ('9. Combustible', costos['combustible']),
                ('10. Flypass', calc.flypass),
                ('11. Peajes', calc.peajes),
                ('12. Urea y/o ACPM', calc.urea_acpm),
                ('13. Cruce Frontera', costos['cruce_frontera']),
                ('14. Hotel', calc.hotel),
                ('15. Comida', calc.comida),
                ('16. Transporte', calc.transporte),
                ('17. Parqueo', costos['parqueo']),
                ('18. Propina/Comisión', calc.propina_comision),
                ('19. Cargue/Descargue', calc.cargue_descargue),
                ('20. Otros', calc.otros),
            ]

            for concepto, monto in detalles_costos:
                ws[f'A{row}'] = concepto
                ws[f'B{row}'] = monto
                ws[f'B{row}'].number_format = '$#,##0'
                row += 1

            row += 1

            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "RESULTADOS"
            cell.font = subheader_font
            cell.fill = subheader_fill
            row += 1

            resultados = [
                ('TOTAL GASTOS', costos['total_gastos']),
                ('LEGALIZACIÓN', costos['legalizacion']),
                ('ANTICIPO', calc.anticipo),
                ('SALDO', costos['saldo']),
                ('PUNTO DE EQUILIBRIO', costos['punto_equilibrio']),
                ('', ''),
                ('VALOR DEL FLETE', calc.valor_flete),
                ('UTILIDAD (UT)', costos['utilidad']),
                ('RENTABILIDAD (%)', costos['rentabilidad']),
                ('', ''),
                ('ANT. EMPRESA (90%)', costos['ant_empresa']),
                ('SALDO EMPRESA', costos['saldo_empresa']),
            ]

            for label, value in resultados:
                if label:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = value
                    ws[f'A{row}'].font = Font(bold=True)
                    if 'TOTAL' in label or 'UTILIDAD' in label:
                        ws[f'A{row}'].fill = total_fill
                        ws[f'B{row}'].fill = total_fill
                        ws[f'B{row}'].font = total_font
                    if label == 'RENTABILIDAD (%)':
                        ws[f'B{row}'].number_format = '#,##0.0"%"'
                    else:
                        ws[f'B{row}'].number_format = '$#,##0'
                row += 1

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generar_excel_totales(df_totales: pd.DataFrame) -> io.BytesIO:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_totales.to_excel(writer, sheet_name='Totales por Flota', index=False)
        output.seek(0)
        return output


# ==================== COMPONENTE DE INPUT CON FORMATO ====================
def input_numero(label, value=0.0, min_value=0.0, step=1000.0, key=None, help=None):
    """Input personalizado que acepta formato colombiano"""
    texto = st.text_input(
        label,
        value=formatear_numero(value) if value > 0 else "",
        key=key,
        help=help,
        placeholder="0"
    )

    numero = limpiar_numero(texto) if texto else 0.0

    if numero > 0:
        st.caption(f"💵 {formatear_numero(numero)} COP")

    return numero


# ==================== MANTENER SESIÓN ACTIVA ====================
def mantener_app_activa():
    if 'ultima_actividad' not in st.session_state:
        st.session_state.ultima_actividad = datetime.now()

    tiempo_inactivo = datetime.now() - st.session_state.ultima_actividad
    segundos_inactivo = int(tiempo_inactivo.total_seconds())

    if segundos_inactivo > 240:
        st.session_state.ultima_actividad = datetime.now()
        st.rerun()

    with st.sidebar:
        st.markdown("---")
        minutos = segundos_inactivo // 60
        segundos = segundos_inactivo % 60
        st.caption(f"⏱️ Sesión activa: {minutos}m {segundos}s")

        if st.button("🔄 Refrescar", key="refresh_manual"):
            st.session_state.ultima_actividad = datetime.now()
            st.rerun()


# ==================== CACHE DE LISTAS BASE (NUEVO v4.6) ====================
@st.cache_data(ttl=30)
def _tractomulas_cached(_db):
    return _db.obtener_tractomulas()


@st.cache_data(ttl=30)
def _conductores_cached(_db):
    return _db.obtener_conductores()


@st.cache_data(ttl=30)
def _rutas_cached(_db):
    return _db.obtener_rutas()


def _refrescar_tractomulas(db):
    _tractomulas_cached.clear()
    st.session_state.tractomulas = db.obtener_tractomulas()


def _refrescar_conductores(db):
    _conductores_cached.clear()
    st.session_state.conductores = db.obtener_conductores()


def _refrescar_rutas(db):
    _rutas_cached.clear()
    st.session_state.rutas = db.obtener_rutas()


# ==================== APLICACIÓN PRINCIPAL ====================
def main():
    st.set_page_config(page_title="Calculadora de Costos Transporte - Colombia 2026", layout="wide")
    mantener_app_activa()

    st.title("🚛 Sistema de Cálculo de Costos para Transporte de Carga")
    st.markdown("**Sistema de Gestión de Flotas y Fletes**")

    if 'datos' not in st.session_state:
        st.session_state.datos = DatosColombia()
    if 'db' not in st.session_state:
        st.session_state.db = DatabaseManager()
    if 'calculadoras' not in st.session_state:
        st.session_state.calculadoras = []

    if 'tractomulas' not in st.session_state:
        st.session_state.tractomulas = _tractomulas_cached(st.session_state.db)
    elif st.session_state.tractomulas and not hasattr(st.session_state.tractomulas[0], 'consumo_urbano'):
        # Sesión con objetos de una versión anterior del código (sin los campos nuevos
        # de consumo por tipo de ruta): forzar recarga fresca desde la base de datos.
        _tractomulas_cached.clear()
        st.session_state.tractomulas = st.session_state.db.obtener_tractomulas()

    if 'conductores' not in st.session_state:
        st.session_state.conductores = _conductores_cached(st.session_state.db)
    if 'rutas' not in st.session_state:
        st.session_state.rutas = _rutas_cached(st.session_state.db)

    datos = st.session_state.datos
    db = st.session_state.db

    with st.sidebar:
        st.header("⚙️ Configuración Global")

        precio_diesel_texto = st.text_input(
            "Precio del Diesel (COP/galón)",
            value=formatear_numero(st.session_state.datos.PRECIO_DIESEL)
        )
        st.session_state.datos.PRECIO_DIESEL = limpiar_numero(precio_diesel_texto)

        st.divider()
        st.subheader("📊 Constantes del Negocio")
        st.caption("Valores fijos según fórmulas ACTUALIZADAS v4.3")
        st.info(f"""
**Nóminas:**
- Admin base: ${formatear_numero(datos.NOMINA_ADMIN_BASE)} / 14
- Conductor/día: ${formatear_numero(datos.NOMINA_CONDUCTOR_DIA)}

**Comisiones conductor:**
- Urbano/Normal: ${formatear_numero(datos.COMISION_URBANO_DIA)} x N° de viajes del día
- Regional: ${formatear_numero(datos.COMISION_REGIONAL)}
- Riohacha: ${formatear_numero(datos.COMISION_RIOACHA)}
- Aguachica: ${formatear_numero(datos.COMISION_AGUACHICA)}
- Frontera: ${formatear_numero(datos.COMISION_FRONTERA)}

**Consumo (km/galón):**
- Ahora es específico por tractomula Y por tipo de ruta (urbano, regional,
  frontera, Aguachica, Riohacha). Se configura en la pestaña "1. Tractomulas".

**Automatización Cliente AGOFER (rutas urbanas):**
- Flete = Peso (kg) x ${formatear_numero(datos.AGOFER_VALOR_POR_KG)} x N° de Viajes
- Cargue/Descargue = ${formatear_numero(datos.AGOFER_CARGUE_DESCARGUE)} x N° de Viajes
- Distancia efectiva = Distancia de la Ruta x N° de Viajes

**Otros:**
- Tecnomecánica/año: ${formatear_numero(datos.TECNOMECANICA_ANUAL)}
- Llantas: ${formatear_numero(datos.LLANTAS_COSTO)}
- Cruce frontera: ${formatear_numero(datos.CRUCE_FRONTERA)}
- Parqueo/día: ${formatear_numero(datos.PARQUEO_DIA)}
- Punto de equilibrio: Valor Flete x {int(datos.PUNTO_EQUILIBRIO_PORCENTAJE*100)}%
        """)

    # ==================== NAVEGACIÓN PERSISTENTE (CORREGIDO v4.6) ====================
    opciones_tabs = [
        "📊 Dashboard",
        "1. Tractomulas",
        "2. Rutas",
        "3. Conductores",
        "4. Cálculo de Viaje",
        "5. Reportes",
        "6. 📂 Trazabilidad",
        "7. Acumulado por Flota",
        "8. 💵 Liquidaciones",
        "9. ⏰ Cuentas Pendientes",
        "10. ⛽ Sobreconsumo",
    ]

    if 'tab_actual' not in st.session_state:
        st.session_state.tab_actual = opciones_tabs[0]

    def _cambiar_tab_actual():
        st.session_state.tab_actual = st.session_state.selector_tab_nav

    st.radio(
        "Navegación",
        opciones_tabs,
        horizontal=True,
        index=opciones_tabs.index(st.session_state.tab_actual),
        key="selector_tab_nav",
        label_visibility="collapsed",
        on_change=_cambiar_tab_actual,
    )
    tab_actual = st.session_state.tab_actual

    # ==================== TAB 0: DASHBOARD ====================
    if tab_actual == opciones_tabs[0]:
        st.header("📊 Dashboard - Resumen de tu Negocio")

        @st.cache_data(ttl=20)
        def _dashboard_data_cached(_db):
            return _db.obtener_dashboard_data()

        dashboard_data = _dashboard_data_cached(db)
        mes_actual = dashboard_data['mes_actual']

        st.subheader(f"📅 Resumen del Mes - {datetime.now().strftime('%B %Y')}")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(
                "💰 Ingresos Totales",
                f"${formatear_numero(mes_actual['total_ingresos'])}",
                help="Total cobrado a clientes este mes"
            )

        with col2:
            st.metric(
                "💸 Gastos Totales",
                f"${formatear_numero(mes_actual['total_gastos'])}",
                help="Total de gastos operativos"
            )

        with col3:
            utilidad = mes_actual['total_utilidad']
            margen_texto = f"{(utilidad/mes_actual['total_gastos']*100) if mes_actual['total_gastos'] > 0 else 0:.1f}% margen"
            st.metric(
                "📈 Utilidad Neta",
                f"${formatear_numero(utilidad)}",
                delta=margen_texto,
                help="Ingresos - Gastos"
            )

        with col4:
            st.metric(
                "🚛 Viajes Realizados",
                f"{mes_actual['total_viajes']}",
                delta=f"{formatear_numero(mes_actual['total_km'])} km",
                help="Total de viajes este mes"
            )

        st.divider()
        st.subheader("💎 Indicadores de Utilidad")

        col1, col2, col3 = st.columns(3)

        with col1:
            ut_bruta = dashboard_data['ut_bruta']
            st.metric("UT BRUTA", f"${formatear_numero(ut_bruta)}",
                      help="Suma de todos los TOTAL CXC (Valor Flete)")

        with col2:
            ut_neta = dashboard_data['ut_neta']
            st.metric("UT NETA", f"${formatear_numero(ut_neta)}",
                      help="Suma de TOTAL UTILIDAD")

        with col3:
            porcentaje_ut = dashboard_data['porcentaje_ut']
            color_ut = "🟢" if porcentaje_ut >= 20 else "🟡" if porcentaje_ut >= 10 else "🔴"
            st.metric("% UT", f"{porcentaje_ut:.1f}%",
                      help="% UT = UT NETA / UT BRUTA")
            st.caption(f"{color_ut} Porcentaje de Utilidad sobre Ingresos Brutos")

        if mes_actual['total_viajes'] > 0:
            col1, col2, col3 = st.columns(3)

            with col1:
                margen = (utilidad / mes_actual['total_ingresos'] * 100) if mes_actual['total_ingresos'] > 0 else 0
                color = "🟢" if margen >= 20 else "🟡" if margen >= 10 else "🔴"
                st.info(f"{color} **Margen de Utilidad:** {margen:.1f}%")

            with col2:
                utilidad_promedio = mes_actual['utilidad_promedio']
                st.info(f"💵 **Utilidad Promedio/Viaje:** ${formatear_numero(utilidad_promedio)}")

            with col3:
                ingreso_promedio = mes_actual['total_ingresos'] / mes_actual['total_viajes']
                st.info(f"🎯 **Ingreso Promedio/Viaje:** ${formatear_numero(ingreso_promedio)}")

        st.divider()

        st.subheader("🚛 Rentabilidad por Tractomula")

        if dashboard_data['por_tractomula']:
            tractomula_df = pd.DataFrame(
                dashboard_data['por_tractomula'],
                columns=['Placa', 'Viajes', 'Gastos', 'Ingresos', 'Utilidad']
            )
            tractomula_df['Margen %'] = ((tractomula_df['Utilidad'] / tractomula_df['Ingresos']) * 100).round(1)
            tractomula_df['Gastos'] = tractomula_df['Gastos'].apply(lambda x: f"${formatear_numero(x)}")
            tractomula_df['Ingresos'] = tractomula_df['Ingresos'].apply(lambda x: f"${formatear_numero(x)}")
            tractomula_df['Utilidad'] = tractomula_df['Utilidad'].apply(lambda x: f"${formatear_numero(x)}")

            st.dataframe(tractomula_df, use_container_width=True, hide_index=True)
        else:
            st.info("No hay datos de tractomulas este mes")

        st.divider()
        st.subheader("Totales Acumulados por Unidad")

        @st.cache_data(ttl=20)
        def _totales_por_placa_cached(_db):
            return _db.obtener_totales_por_placa()

        placa_seleccionada = st.selectbox("Selecciona una placa", sorted(PLACA_CONDUCTOR.keys()))
        if placa_seleccionada:
            df_totales = _totales_por_placa_cached(db)
            if not df_totales.empty:
                df_placa = df_totales[df_totales['placa'] == placa_seleccionada]
                if not df_placa.empty:
                    row = df_placa.iloc[0]
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total CXC", f"${formatear_numero(row['total_cxc'])}")
                        st.metric("Total Gastos", f"${formatear_numero(row['total_gastos'])}")
                    with col2:
                        st.metric("Total UT", f"${formatear_numero(row['total_ut'])}")
                        st.metric("Rentabilidad", f"{row['total_rentabilidad']:.1f}%")
                    with col3:
                        st.metric("Punto Equilibrio (40% CXC)", f"${formatear_numero(row['total_punto_equilibrio'])}")
                        st.metric("Saldo Total", f"${formatear_numero(row['total_saldo'])}")
                else:
                    st.info("No hay datos para esta placa")
            else:
                st.info("No hay datos para esta placa")

        st.subheader("Comparativa entre Unidades")
        df_totales = _totales_por_placa_cached(db)
        if not df_totales.empty:
            fig_utilidad = px.bar(df_totales, x='placa', y='total_ut', title="Utilidad Total por Unidad")
            st.plotly_chart(fig_utilidad)
            fig_gastos = px.bar(df_totales, x='placa', y='total_gastos', title="Gastos Totales por Unidad")
            st.plotly_chart(fig_gastos)
        else:
            st.info("No hay datos para comparar")

    # ==================== TAB 1: TRACTOMULAS ====================
    if tab_actual == opciones_tabs[1]:
        st.header("Tus Tractomulas")
        st.caption("💡 El rendimiento (km/galón) de una misma tractomula cambia según el tipo de ruta "
                   "(por ejemplo: 5 km/gal en urbano, 7 km/gal en regional). Define un valor para cada tipo; "
                   "el sistema elige automáticamente el correcto según la ruta del viaje.")

        placas_opciones = ['(Escribir nueva)'] + sorted(PLACA_CONDUCTOR.keys())
        placa_seleccion = st.selectbox("Placa", placas_opciones, key="tractomula_placa_sel")
        if placa_seleccion == '(Escribir nueva)':
            placa_ingresada = st.text_input("Placa manual", key="tractomula_placa_manual")
            placa = placa_ingresada.strip().upper()
        else:
            placa = placa_seleccion

        with st.form(key="form_tractomula"):
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**Placa seleccionada:** {placa or '(sin definir)'}")
                tipo = st.selectbox("Tipo", ["Sencilla", "Dobletroque", "Minimula", "Otro"])
            with col2:
                st.write("**Consumo por tipo de ruta (km/galón)**")

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                consumo_urbano = st.number_input("Urbano", min_value=0.0, value=5.0, step=0.1, key="nueva_consumo_urbano")
            with col2:
                consumo_regional = st.number_input("Regional", min_value=0.0, value=5.0, step=0.1, key="nueva_consumo_regional")
            with col3:
                consumo_frontera = st.number_input("Frontera", min_value=0.0, value=5.0, step=0.1, key="nueva_consumo_frontera")
            with col4:
                consumo_aguachica = st.number_input("Aguachica", min_value=0.0, value=5.0, step=0.1, key="nueva_consumo_aguachica")
            with col5:
                consumo_riohacha = st.number_input("Riohacha", min_value=0.0, value=5.0, step=0.1, key="nueva_consumo_riohacha")

            submit = st.form_submit_button("Agregar Tractomula")
            if submit and placa:
                tractomula = Tractomula(
                    placa=placa, consumo_km_galon=consumo_urbano, tipo=tipo,
                    consumo_urbano=consumo_urbano, consumo_regional=consumo_regional,
                    consumo_frontera=consumo_frontera, consumo_aguachica=consumo_aguachica,
                    consumo_riohacha=consumo_riohacha
                )
                if db.guardar_tractomula(tractomula):
                    _refrescar_tractomulas(db)
                    st.success(f"✅ Tractomula {placa} guardada!")
                    st.rerun()
                else:
                    st.error(f"❌ La placa {placa} ya existe")
            elif submit and not placa:
                st.error("⚠️ Debes indicar una placa (selecciónala o escríbela arriba).")

        if st.session_state.tractomulas:
            st.subheader("Tractomulas Registradas")
            for idx, t in enumerate(st.session_state.tractomulas):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(
                        f"**{t.placa}** ({t.tipo}) — "
                        f"Urbano: {t.consumo_urbano} km/gal | Regional: {t.consumo_regional} km/gal | "
                        f"Frontera: {t.consumo_frontera} km/gal | Aguachica: {t.consumo_aguachica} km/gal | "
                        f"Riohacha: {t.consumo_riohacha} km/gal"
                    )
                with col2:
                    if st.button("🗑️", key=f"eliminar_tractomula_{idx}"):
                        db.eliminar_tractomula(t.placa)
                        _refrescar_tractomulas(db)
                        st.success(f"Tractomula {t.placa} eliminada")
                        st.rerun()

                with st.expander(f"✏️ Editar consumos de {t.placa}"):
                    with st.form(key=f"form_editar_tractomula_{idx}"):
                        ecol1, ecol2, ecol3, ecol4, ecol5 = st.columns(5)
                        with ecol1:
                            e_urbano = st.number_input("Urbano", min_value=0.0, value=float(t.consumo_urbano), step=0.1, key=f"edit_urbano_{idx}")
                        with ecol2:
                            e_regional = st.number_input("Regional", min_value=0.0, value=float(t.consumo_regional), step=0.1, key=f"edit_regional_{idx}")
                        with ecol3:
                            e_frontera = st.number_input("Frontera", min_value=0.0, value=float(t.consumo_frontera), step=0.1, key=f"edit_frontera_{idx}")
                        with ecol4:
                            e_aguachica = st.number_input("Aguachica", min_value=0.0, value=float(t.consumo_aguachica), step=0.1, key=f"edit_aguachica_{idx}")
                        with ecol5:
                            e_riohacha = st.number_input("Riohacha", min_value=0.0, value=float(t.consumo_riohacha), step=0.1, key=f"edit_riohacha_{idx}")
                        tipos_disponibles = ["Sencilla", "Dobletroque", "Minimula", "Otro"]
                        e_tipo = st.selectbox(
                            "Tipo", tipos_disponibles,
                            index=tipos_disponibles.index(t.tipo) if t.tipo in tipos_disponibles else 0,
                            key=f"edit_tipo_{idx}"
                        )
                        if st.form_submit_button("💾 Guardar cambios"):
                            t_editada = Tractomula(
                                placa=t.placa, consumo_km_galon=e_urbano, tipo=e_tipo,
                                consumo_urbano=e_urbano, consumo_regional=e_regional,
                                consumo_frontera=e_frontera, consumo_aguachica=e_aguachica,
                                consumo_riohacha=e_riohacha
                            )
                            if db.actualizar_tractomula(t_editada):
                                _refrescar_tractomulas(db)
                                st.success(f"✅ Consumos de {t.placa} actualizados")
                                st.rerun()

    # ==================== TAB 2: RUTAS ====================
    if tab_actual == opciones_tabs[2]:
        st.header("Tus Rutas")
        st.caption("💡 Los valores por defecto se autocompletan en cada viaje nuevo con esta ruta (puedes editarlos si cambian).")
        with st.form(key="form_ruta"):
            col1, col2 = st.columns(2)
            with col1:
                origen = st.text_input("Origen")
                destino = st.text_input("Destino")
                distancia_km = st.number_input("Distancia (km)", min_value=0.0)
                ida_vuelta = st.checkbox("Ida y vuelta")
            with col2:
                es_frontera = st.checkbox("¿Es ruta a frontera?", help="Comisión conductor: $565.000")
                es_regional = st.checkbox("¿Es regional?", help="Comisión conductor: $200.000")
                es_aguachica = st.checkbox("¿Es para Aguachica?", help="Comisión conductor: $360.000")
                es_riohacha = st.checkbox("¿Es para Riohacha?", help="Comisión conductor: $350.000")

            st.divider()
            st.markdown("**Valores por defecto de gastos variables (opcional, se autocompletan en cada viaje)**")
            col1, col2, col3 = st.columns(3)
            with col1:
                default_flypass_texto = st.text_input("Flypass default (COP)", value="", placeholder="0")
                default_flypass = limpiar_numero(default_flypass_texto)
                if default_flypass > 0:
                    st.caption(f"💵 {formatear_numero(default_flypass)}")

                default_peajes_texto = st.text_input("Peajes default (COP)", value="", placeholder="0")
                default_peajes = limpiar_numero(default_peajes_texto)
                if default_peajes > 0:
                    st.caption(f"💵 {formatear_numero(default_peajes)}")

                default_urea_acpm_texto = st.text_input("Urea y/o ACPM default (COP)", value="", placeholder="0")
                default_urea_acpm = limpiar_numero(default_urea_acpm_texto)
                if default_urea_acpm > 0:
                    st.caption(f"💵 {formatear_numero(default_urea_acpm)}")
            with col2:
                default_hotel_texto = st.text_input("Hotel default (COP)", value="", placeholder="0")
                default_hotel = limpiar_numero(default_hotel_texto)
                if default_hotel > 0:
                    st.caption(f"💵 {formatear_numero(default_hotel)}")

                default_comida_texto = st.text_input("Comida default (COP)", value="", placeholder="0")
                default_comida = limpiar_numero(default_comida_texto)
                if default_comida > 0:
                    st.caption(f"💵 {formatear_numero(default_comida)}")

                default_transporte_texto = st.text_input("Transporte default (COP)", value="", placeholder="0")
                default_transporte = limpiar_numero(default_transporte_texto)
                if default_transporte > 0:
                    st.caption(f"💵 {formatear_numero(default_transporte)}")
            with col3:
                default_propina_comision_texto = st.text_input("Propina/Comisión default (COP)", value="", placeholder="0")
                default_propina_comision = limpiar_numero(default_propina_comision_texto)
                if default_propina_comision > 0:
                    st.caption(f"💵 {formatear_numero(default_propina_comision)}")

                default_cargue_descargue_texto = st.text_input(
                    "Cargue/Descargue-Amarre default (COP)", value="", placeholder="0",
                    help="Para rutas urbanas con Cliente = AGOFER, este valor se autocalcula (30.000 x N° Viajes) en el momento del viaje y no depende de este default."
                )
                default_cargue_descargue = limpiar_numero(default_cargue_descargue_texto)
                if default_cargue_descargue > 0:
                    st.caption(f"💵 {formatear_numero(default_cargue_descargue)}")

                default_otros_texto = st.text_input("Otros default (COP)", value="", placeholder="0")
                default_otros = limpiar_numero(default_otros_texto)
                if default_otros > 0:
                    st.caption(f"💵 {formatear_numero(default_otros)}")

            submit = st.form_submit_button("Agregar Ruta")
            if submit and origen and destino:
                if ida_vuelta:
                    distancia_km *= 2
                    destino = f"{destino} (ida y vuelta)"
                ruta = Ruta(
                    origen=origen, destino=destino, distancia_km=distancia_km,
                    es_frontera=es_frontera, es_regional=es_regional,
                    es_aguachica=es_aguachica, es_riohacha=es_riohacha,
                    default_flypass=default_flypass, default_peajes=default_peajes,
                    default_urea_acpm=default_urea_acpm, default_hotel=default_hotel,
                    default_comida=default_comida, default_transporte=default_transporte,
                    default_propina_comision=default_propina_comision,
                    default_cargue_descargue=default_cargue_descargue, default_otros=default_otros
                )
                ruta_id = db.guardar_ruta(ruta)
                _refrescar_rutas(db)
                st.success(f"✅ Ruta {origen} → {destino} guardada! (ID: {ruta_id})")
                st.rerun()

        if st.session_state.rutas:
            st.subheader("Rutas Registradas")
            rutas_con_id = db.obtener_rutas_con_id()

            for ruta_data in rutas_con_id:
                ruta_id = ruta_data[0]
                origen = ruta_data[1]
                destino = ruta_data[2]
                dist = ruta_data[3]
                es_front = bool(ruta_data[4])
                es_reg = bool(ruta_data[5])
                es_agua = bool(ruta_data[6])
                es_rioh = bool(ruta_data[7])

                col1, col2 = st.columns([4, 1])
                with col1:
                    tags = []
                    if es_front:
                        tags.append("🌐 FRONTERA ($565k)")
                    if es_reg:
                        tags.append("📍 REGIONAL ($200k)")
                    if es_agua:
                        tags.append("🏙️ AGUACHICA ($360k)")
                    if es_rioh:
                        tags.append("🏖️ RIOHACHA ($350k)")

                    tags_str = " ".join(tags) if tags else "🚛 URBANO ($120k x N° viajes)"
                    st.write(f"**{origen}** → **{destino}** ({formatear_numero(dist)} km) {tags_str}")
                with col2:
                    if st.button("🗑️", key=f"eliminar_ruta_{ruta_id}"):
                        db.eliminar_ruta(ruta_id)
                        _refrescar_rutas(db)
                        st.success("Ruta eliminada")
                        st.rerun()

    # ==================== TAB 3: CONDUCTORES ====================
    if tab_actual == opciones_tabs[3]:
        st.header("Tus Conductores")

        if 'conductores_cedulas' not in st.session_state:
            st.session_state.conductores_cedulas = {
                "JOSE ORTEGA PEREZ": "987654321",
                "ISAIAS VESGA": "789123456",
                "FLAVIO ROSENDO MALTE TUTALCHA": "321654987",
                "SLITH JOSE ORTEGA PACHECO": "654987321",
                "ABRAHAM SEGUNDO ALVAREZ VALLE": "147258369",
                "RAMON TAFUR HERNANDEZ": "369258147",
                "PEDRO VILLAMIL": "258369147",
                "JESUS DAVID MONTE MOSQUERA": "951753486",
                "CHRISTIAN MARTINEZ NAVARRO": "486159753",
                "YEIMI DUQUE ZULUAGA": "753486159",
                "JULIAN CALETH CORONADO": "159753486",
                "CARLOS TAFUR": "357159486",
                "EDUARDO RAFAEL OLIVARES ALCAZAR": "486357159"
            }

        nombres_opciones = ['(Escribir nuevo)'] + sorted([n for n in PLACA_CONDUCTOR.values() if n])
        nombre_seleccion = st.selectbox("Nombre", nombres_opciones, key="conductor_nombre_sel")
        if nombre_seleccion == '(Escribir nuevo)':
            nombre = st.text_input("Nombre manual", key="conductor_nombre_manual")
            cedula_auto = ""
        else:
            nombre = nombre_seleccion
            cedula_auto = st.session_state.conductores_cedulas.get(nombre, "")

        with st.form(key="form_conductor"):
            st.write(f"**Nombre seleccionado:** {nombre or '(sin definir)'}")
            if cedula_auto:
                cedula = st.text_input("Cédula", value=cedula_auto)
                st.info("📋 Cédula encontrada automáticamente")
            else:
                cedula = st.text_input("Cédula")

            submit = st.form_submit_button("Agregar Conductor")
            if submit and nombre and cedula:
                conductor = Conductor(nombre, cedula)
                if db.guardar_conductor(conductor):
                    _refrescar_conductores(db)
                    st.session_state.conductores_cedulas[nombre] = cedula
                    st.success(f"✅ Conductor {nombre} guardado!")
                    st.rerun()
                else:
                    st.error(f"❌ El conductor {nombre} ya existe")
            elif submit and not nombre:
                st.error("⚠️ Debes indicar un nombre (selecciónalo o escríbelo arriba).")

        if st.session_state.conductores:
            st.subheader("Conductores Registrados")
            for idx, c in enumerate(st.session_state.conductores):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(f"**{c.nombre}** - Cédula: {c.cedula}")
                with col2:
                    if st.button("🗑️", key=f"eliminar_conductor_{idx}"):
                        db.eliminar_conductor(c.nombre)
                        _refrescar_conductores(db)
                        st.success(f"Conductor {c.nombre} eliminado")
                        st.rerun()

    # ==================== TAB 4: CÁLCULO DE VIAJE ====================
    if tab_actual == opciones_tabs[4]:
        st.header("Realizar Cálculo de Viaje")

        # ---------------- NUEVO v4.8: Checkbox de Día Vacío ----------------
        dia_vacio = st.checkbox(
            "🚫 Día vacío (el carro NO hizo viaje este día)",
            key="check_dia_vacio",
            help="Actívalo cuando la tractomula no rodó ese día. Solo se guarda un registro de trazabilidad "
                 "(fecha, placa, conductor y motivo). NO se calcula ni se guarda ningún costo, gasto, flete ni utilidad."
        )

        if dia_vacio:
            st.info("📭 Se registrará únicamente que este carro NO hizo viaje ese día. No se calculará ningún costo, gasto ni utilidad.")

            placas_disponibles_vacio = (
                [t.placa for t in st.session_state.tractomulas]
                if st.session_state.tractomulas else sorted(PLACA_CONDUCTOR.keys())
            )
            conductores_disponibles_vacio = (
                [c.nombre for c in st.session_state.conductores]
                if st.session_state.conductores else []
            )

            col1, col2 = st.columns(2)
            with col1:
                placa_vacio = st.selectbox("Placa", placas_disponibles_vacio, key="vacio_placa")
                conductor_vacio = st.selectbox(
                    "Conductor (opcional)",
                    ["(Ninguno)"] + conductores_disponibles_vacio,
                    key="vacio_conductor"
                )
            with col2:
                fecha_vacio = st.date_input("Fecha", value=datetime.now().date(), key="vacio_fecha")
                motivo_vacio = st.text_input(
                    "Motivo (opcional)",
                    placeholder="Ej: en taller, sin carga, descanso",
                    key="vacio_motivo"
                )

            obs_vacio = st.text_area("Observaciones (opcional)", key="vacio_obs")

            if st.button("💾 Registrar Día Vacío", type="primary", key="btn_guardar_vacio"):
                cond_val = "" if conductor_vacio == "(Ninguno)" else conductor_vacio
                dia_id = db.guardar_dia_sin_viaje(fecha_vacio, placa_vacio, cond_val, motivo_vacio, obs_vacio)
                if dia_id:
                    st.success(f"✅ Registrado: {placa_vacio} NO hizo viaje el {fecha_vacio.strftime('%Y-%m-%d')} (ID: {dia_id})")
                else:
                    st.error("❌ Error al guardar el registro")

        else:
            if not (st.session_state.tractomulas and st.session_state.conductores and st.session_state.rutas):
                st.warning("⚠️ Primero agrega al menos una tractomula, un conductor y una ruta.")
            else:
                st.subheader("📋 Selección Básica")
                col1, col2 = st.columns(2)
                with col1:
                    tractomula_selec = st.selectbox("Selecciona Tractomula", [t.placa for t in st.session_state.tractomulas], key="sel_tractomula")
                    tractomula_obj = next(t for t in st.session_state.tractomulas if t.placa == tractomula_selec)

                    conductores = [c.nombre for c in st.session_state.conductores]
                    conductor_asignado = PLACA_CONDUCTOR.get(tractomula_selec)
                    conductor_index = conductores.index(conductor_asignado) if conductor_asignado in conductores else 0
                    conductor_selec = st.selectbox("Selecciona Conductor", conductores, index=conductor_index, key="sel_conductor")
                    conductor_obj = next(c for c in st.session_state.conductores if c.nombre == conductor_selec)

                with col2:
                    ruta_selec = st.selectbox("Selecciona Ruta", [f"{r.origen} → {r.destino}" for r in st.session_state.rutas], key="sel_ruta")
                    ruta_obj = next(r for r in st.session_state.rutas if f"{r.origen} → {r.destino}" == ruta_selec)
                    dias_viaje = st.number_input("Días del viaje", min_value=1, value=1, step=1, key="sel_dias")
                    numero_viajes = st.number_input(
                        "🚛 Número de viajes",
                        min_value=1, value=1, step=1, key="sel_numero_viajes",
                        help="Cuántos viajes hizo el conductor este día. Afecta la Comisión Conductor Urbano/Normal, y para el cliente AGOFER en rutas urbanas también afecta el Flete, el Cargue/Descargue y la distancia recorrida."
                    )
                    fecha_viaje = st.date_input(
                        "📅 Fecha del viaje",
                        value=datetime.now().date(),
                        help="Fecha real en que ocurrió/ocurrirá el viaje (no la fecha en que lo registras). Se usa para todos los filtros de fecha del sistema.",
                        key="sel_fecha_viaje"
                    )
                    cliente_viaje = st.text_input(
                        "🏢 Cliente",
                        value="",
                        placeholder="Nombre de la empresa o persona que contrató el flete",
                        help="Para poder buscar y ver la trazabilidad de todos los viajes hechos para este cliente. Escribe 'AGOFER' para activar la automatización de Flete/Cargue-Descargue/Distancia en rutas urbanas.",
                        key="sel_cliente"
                    )
                    peso_texto = st.text_input(
                        "⚖️ Peso transportado (kg)",
                        value="",
                        placeholder="Ejemplo: 30.000",
                        help="Peso de la carga en kilogramos (NO en toneladas). Se usa para calcular automáticamente el Flete cuando el cliente es AGOFER en rutas urbanas (Flete = Peso x 27.500 x N° de Viajes).",
                        key="sel_peso"
                    )
                    peso = limpiar_numero(peso_texto)
                    if peso > 0:
                        st.caption(f"⚖️ {formatear_numero(peso)} kg")

                aplica_agofer = ruta_obj.es_urbana and es_cliente_agofer(cliente_viaje)
                flete_sugerido_agofer = peso * datos.AGOFER_VALOR_POR_KG * numero_viajes if aplica_agofer else 0.0
                cargue_sugerido_agofer = datos.AGOFER_CARGUE_DESCARGUE * numero_viajes if aplica_agofer else 0.0
                distancia_sugerida_agofer = ruta_obj.distancia_km * numero_viajes if aplica_agofer else ruta_obj.distancia_km

                if aplica_agofer:
                    st.success(
                        f"🤖 **Automatización AGOFER activa** (ruta urbana + cliente AGOFER): "
                        f"Flete sugerido = ${formatear_numero(flete_sugerido_agofer)} "
                        f"(Peso {formatear_numero(peso)} kg x ${formatear_numero(datos.AGOFER_VALOR_POR_KG)} x {numero_viajes} viaje(s)) · "
                        f"Cargue/Descargue sugerido = ${formatear_numero(cargue_sugerido_agofer)} "
                        f"(${formatear_numero(datos.AGOFER_CARGUE_DESCARGUE)} x {numero_viajes} viaje(s)) · "
                        f"Distancia efectiva del día = {formatear_numero(distancia_sugerida_agofer)} km "
                        f"({formatear_numero(ruta_obj.distancia_km)} km x {numero_viajes} viaje(s)). "
                        f"Estos valores ya vienen precargados abajo y puedes editarlos si el viaje es distinto."
                    )

                # ---------------- Consumo km/galón según tipo de ruta ----------------
                if ruta_obj.es_aguachica:
                    _tipo_ruta_label = "Aguachica"
                    _consumo_previo = tractomula_obj.consumo_aguachica
                elif ruta_obj.es_riohacha:
                    _tipo_ruta_label = "Riohacha"
                    _consumo_previo = tractomula_obj.consumo_riohacha
                elif ruta_obj.es_regional:
                    _tipo_ruta_label = "Regional"
                    _consumo_previo = tractomula_obj.consumo_regional
                elif ruta_obj.es_frontera:
                    _tipo_ruta_label = "Frontera"
                    _consumo_previo = tractomula_obj.consumo_frontera
                else:
                    _tipo_ruta_label = "Urbano"
                    _consumo_previo = tractomula_obj.consumo_urbano
                _consumo_previo = _consumo_previo if _consumo_previo and _consumo_previo > 0 else tractomula_obj.consumo_km_galon
                st.caption(f"⛽ Consumo que se usará para {tractomula_obj.placa} en esta ruta ({_tipo_ruta_label}): **{_consumo_previo} km/galón**. "
                           f"Se ajusta en la pestaña '1. Tractomulas'.")

                st.caption("💡 Los campos de gastos variables abajo ya vienen precargados con los valores por defecto de esta ruta (o con los calculados automáticamente para AGOFER). Puedes editarlos si el viaje tuvo un valor distinto.")

                with st.form(key="form_calculo"):
                    st.divider()
                    st.subheader("📊 Parámetros del Viaje")

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        es_frontera = st.checkbox("¿Es viaje a frontera?", value=ruta_obj.es_frontera,
                                                    help="Afecta Comisión Conductor y Cruce Frontera")
                        hubo_parqueo = st.checkbox("¿Hubo parqueo?", value=False)
                        hubo_anticipo_empresa = st.checkbox("¿Hubo anticipo empresa?", value=False,
                                                           help="Activa ANTICIPO EMPRESA = VALOR FLETE × 0.90")

                    with col2:
                        flypass_texto = st.text_input("Flypass (COP)", value=formatear_numero(ruta_obj.default_flypass) if ruta_obj.default_flypass > 0 else "", placeholder="0")
                        flypass = limpiar_numero(flypass_texto)
                        if flypass > 0:
                            st.caption(f"💵 {formatear_numero(flypass)}")

                        peajes_texto = st.text_input("Peajes (COP)", value=formatear_numero(ruta_obj.default_peajes) if ruta_obj.default_peajes > 0 else "", placeholder="0")
                        peajes = limpiar_numero(peajes_texto)
                        if peajes > 0:
                            st.caption(f"💵 {formatear_numero(peajes)}")

                        urea_acpm_texto = st.text_input("Urea y/o ACPM (COP)", value=formatear_numero(ruta_obj.default_urea_acpm) if ruta_obj.default_urea_acpm > 0 else "", placeholder="0")
                        urea_acpm = limpiar_numero(urea_acpm_texto)
                        if urea_acpm > 0:
                            st.caption(f"💵 {formatear_numero(urea_acpm)}")

                    with col3:
                        hotel_texto = st.text_input("Hotel (COP)", value=formatear_numero(ruta_obj.default_hotel) if ruta_obj.default_hotel > 0 else "", placeholder="0")
                        hotel = limpiar_numero(hotel_texto)
                        if hotel > 0:
                            st.caption(f"💵 {formatear_numero(hotel)}")

                        comida_texto = st.text_input("Comida (COP)", value=formatear_numero(ruta_obj.default_comida) if ruta_obj.default_comida > 0 else "", placeholder="0")
                        comida = limpiar_numero(comida_texto)
                        if comida > 0:
                            st.caption(f"💵 {formatear_numero(comida)}")

                        transporte_texto = st.text_input("Transporte (COP)", value=formatear_numero(ruta_obj.default_transporte) if ruta_obj.default_transporte > 0 else "", placeholder="0")
                        transporte = limpiar_numero(transporte_texto)
                        if transporte > 0:
                            st.caption(f"💵 {formatear_numero(transporte)}")

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        propina_texto = st.text_input("Propina/Comisión (COP)", value=formatear_numero(ruta_obj.default_propina_comision) if ruta_obj.default_propina_comision > 0 else "", placeholder="0")
                        propina_comision = limpiar_numero(propina_texto)
                        if propina_comision > 0:
                            st.caption(f"💵 {formatear_numero(propina_comision)}")
                    with col2:
                        valor_default_cargue = cargue_sugerido_agofer if aplica_agofer else ruta_obj.default_cargue_descargue
                        cargue_texto = st.text_input(
                            "Cargue/Descargue - Amarre/Desamarre (COP)",
                            value=formatear_numero(valor_default_cargue) if valor_default_cargue > 0 else "",
                            placeholder="0",
                            help="Para cliente AGOFER en rutas urbanas se autocalcula: 30.000 x Número de Viajes. Editable si el viaje es distinto."
                        )
                        cargue_descargue = limpiar_numero(cargue_texto)
                        if cargue_descargue > 0:
                            st.caption(f"💵 {formatear_numero(cargue_descargue)}")
                    with col3:
                        otros_texto = st.text_input("Otros - Engrase, Lavada, Policía (COP)", value=formatear_numero(ruta_obj.default_otros) if ruta_obj.default_otros > 0 else "", placeholder="0")
                        otros = limpiar_numero(otros_texto)
                        if otros > 0:
                            st.caption(f"💵 {formatear_numero(otros)}")

                    st.divider()
                    col1, col2 = st.columns(2)
                    with col1:
                        anticipo_texto = st.text_input("Anticipo (COP)", value="", placeholder="0",
                                                      help="Anticipo entregado al conductor")
                        anticipo = limpiar_numero(anticipo_texto)
                        if anticipo > 0:
                            st.caption(f"💵 {formatear_numero(anticipo)}")

                    st.divider()
                    st.subheader("💰 Valor del Flete")

                    valor_flete_default = flete_sugerido_agofer if aplica_agofer else 0.0
                    valor_flete_texto = st.text_input(
                        "💰 Valor del Flete Cobrado al Cliente (COP)",
                        value=formatear_numero(valor_flete_default) if valor_flete_default > 0 else "",
                        placeholder="Ejemplo: 5.000.000",
                        help="¿Cuánto VAS A COBRAR o YA COBRASTE por este viaje? Para cliente AGOFER en rutas urbanas se autocalcula: Peso x 27.500 x N° de Viajes. Editable si el viaje es distinto."
                    )
                    valor_flete = limpiar_numero(valor_flete_texto)

                    if valor_flete > 0:
                        st.success(f"✅ Flete: ${formatear_numero(valor_flete)} COP")

                    if valor_flete > 0:
                        calc_preview = CalculadoraCostos(
                            tractomula_obj, conductor_obj, ruta_obj,
                            dias_viaje, numero_viajes, es_frontera, hubo_parqueo,
                            flypass, peajes, urea_acpm, hotel, comida, transporte,
                            propina_comision, cargue_descargue, otros,
                            valor_flete, anticipo, hubo_anticipo_empresa, datos,
                            peso=peso, cliente=cliente_viaje
                        )
                        costos_preview = calc_preview.calcular_costos_totales()

                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Gastos", f"${formatear_numero(costos_preview['total_gastos'])}")
                        with col2:
                            st.metric("Utilidad", f"${formatear_numero(costos_preview['utilidad'])}",
                                      delta=f"{costos_preview['rentabilidad']:.1f}%")
                        with col3:
                            st.metric("Saldo", f"${formatear_numero(costos_preview['saldo'])}",
                                      help="ANTICIPO - LEGALIZACIÓN")
                        with col4:
                            st.metric("Punto Equilibrio (40% Flete)", f"${formatear_numero(costos_preview['punto_equilibrio'])}")

                        st.info(f"🚛 **Comisión Conductor calculada:** ${formatear_numero(costos_preview['comision_conductor'])} "
                                f"(según la ruta: fija, o $120.000 x {numero_viajes} viaje(s) si es urbana) · "
                                f"**Distancia efectiva usada en combustible/llantas/aceite:** {formatear_numero(calc_preview.distancia_efectiva)} km · "
                                f"**Consumo usado:** {calc_preview.obtener_consumo_km_galon()} km/galón ({_tipo_ruta_label})")

                    observaciones = st.text_area("Observaciones (opcional)", placeholder="Notas sobre este viaje...")

                    col_btn1, col_btn2 = st.columns(2)
                    with col_btn1:
                        calcular = st.form_submit_button("📊 Calcular Costos", type="primary")
                    with col_btn2:
                        guardar = st.form_submit_button("💾 Calcular y Guardar", type="secondary")

                    if calcular or guardar:
                        if valor_flete <= 0:
                            st.error("⚠️ Debes ingresar el Valor del Flete para continuar")
                        else:
                            calculadora = CalculadoraCostos(
                                tractomula_obj, conductor_obj, ruta_obj,
                                dias_viaje, numero_viajes, es_frontera, hubo_parqueo,
                                flypass, peajes, urea_acpm, hotel, comida, transporte,
                                propina_comision, cargue_descargue, otros,
                                valor_flete, anticipo, hubo_anticipo_empresa, datos,
                                peso=peso, cliente=cliente_viaje
                            )
                            st.session_state.calculadoras.append(calculadora)

                            if guardar:
                                viaje_id = db.guardar_viaje(calculadora, fecha_viaje, observaciones, cliente_viaje)
                                if viaje_id:
                                    costos = calculadora.calcular_costos_totales()
                                    utilidad = costos.get('utilidad', 0)
                                    if utilidad >= 0:
                                        st.success(f"""
                                        ✅ **Viaje guardado exitosamente (ID: {viaje_id})**

                                        - Fecha del Viaje: {fecha_viaje.strftime('%Y-%m-%d')}
                                        - Número de Viajes: {numero_viajes}
                                        - Peso: {formatear_numero(peso)} kg
                                        - Distancia efectiva: {formatear_numero(calculadora.distancia_efectiva)} km
                                        - Total Gastos: ${formatear_numero(costos['total_gastos'])}
                                        - Valor Flete: ${formatear_numero(calculadora.valor_flete)}
                                        - **Utilidad: ${formatear_numero(utilidad)}**
                                        - **Rentabilidad: {costos['rentabilidad']:.1f}%**
                                        - **Saldo: ${formatear_numero(costos['saldo'])}**
                                        """)
                                    else:
                                        st.error(f"""
                                        ⚠️ **Viaje guardado (ID: {viaje_id}) - PÉRDIDA DETECTADA**

                                        - Fecha del Viaje: {fecha_viaje.strftime('%Y-%m-%d')}
                                        - Número de Viajes: {numero_viajes}
                                        - Peso: {formatear_numero(peso)} kg
                                        - Distancia efectiva: {formatear_numero(calculadora.distancia_efectiva)} km
                                        - Total Gastos: ${formatear_numero(costos['total_gastos'])}
                                        - Valor Flete: ${formatear_numero(calculadora.valor_flete)}
                                        - **Pérdida: ${formatear_numero(utilidad)}**
                                        - **Rentabilidad: {costos['rentabilidad']:.1f}%**

                                        ⚠️ Este viaje NO fue rentable.
                                        """)
                                else:
                                    st.error("❌ Error al guardar el viaje en la base de datos.")
                            else:
                                st.success("✅ Cálculo completado! Ve a la pestaña de Reportes.")

    # ==================== TAB 5: REPORTES ====================
    if tab_actual == opciones_tabs[5]:
        st.header("📄 Reportes y Descargas")
        if st.session_state.calculadoras:
            for idx, calc in enumerate(st.session_state.calculadoras, 1):
                st.subheader(f"Reporte {idx}: {calc.ruta.origen} → {calc.ruta.destino}")
                st.text(GeneradorReportes.generar_reporte_texto(calc))

            excel_data = GeneradorReportes.generar_excel(st.session_state.calculadoras)
            ultimo = st.session_state.calculadoras[-1]
            conductor_nombre = ultimo.conductor.nombre.strip()
            placa = ultimo.tractomula.placa.strip()
            fecha_archivo = datetime.now().strftime('%d-%m-%Y')
            nombre_archivo = f"{conductor_nombre} {placa} {fecha_archivo}.xlsx"
            nombre_archivo = re.sub(r'[\\/:*?"<>|]', '-', nombre_archivo)

            st.download_button(
                label="📥 Descargar Reporte Completo en Excel",
                data=excel_data,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            if st.button("🗑️ Limpiar reportes temporales"):
                st.session_state.calculadoras = []
                st.rerun()
        else:
            st.info("Realiza al menos un cálculo en la pestaña anterior para ver reportes.")

    # ==================== TAB 6: TRAZABILIDAD ====================
    if tab_actual == opciones_tabs[6]:
        st.header("📂 Trazabilidad de Viajes")
        st.markdown("Historial completo de todos los viajes guardados en el sistema.")

        with st.expander("🔍 Filtros de Búsqueda", expanded=True):
            col1, col2, col3 = st.columns(3)

            with col1:
                fecha_inicio = st.date_input("Fecha desde", value=None)
                placa_filtro = st.selectbox("Placa", ["Todas"] + sorted(PLACA_CONDUCTOR.keys()))

            with col2:
                fecha_fin = st.date_input("Fecha hasta", value=None)
                conductor_filtro = st.text_input("Conductor (nombre)")
            with col3:
                origen_filtro = st.text_input("Origen")
                destino_filtro = st.text_input("Destino")

            cliente_filtro = st.text_input("🏢 Cliente")

            buscar = st.button("🔍 Buscar", type="primary")

        primera_vez = 'ultima_busqueda' not in st.session_state
        if buscar or primera_vez:
            fecha_ini = fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else None
            fecha_fi = fecha_fin.strftime('%Y-%m-%d') if fecha_fin else None
            placa_f = None if placa_filtro == "Todas" else placa_filtro
            conductor_f = conductor_filtro if conductor_filtro else None
            origen_f = origen_filtro if origen_filtro else None
            destino_f = destino_filtro if destino_filtro else None
            cliente_f = cliente_filtro if cliente_filtro else None

            df_viajes = db.buscar_viajes(fecha_ini, fecha_fi, placa_f, conductor_f, origen_f, destino_f, cliente_f)
            st.session_state.ultima_busqueda = df_viajes
        else:
            df_viajes = st.session_state.ultima_busqueda

        if df_viajes.empty:
            st.info("No se encontraron viajes con los filtros aplicados.")
        else:
            st.success(f"Se encontraron {len(df_viajes)} viajes")

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Viajes", int(df_viajes['numero_viajes'].fillna(1).sum()))
                st.caption(f"({len(df_viajes)} registros)")
            with col2:
                st.metric("Kilómetros", f"{formatear_numero(df_viajes['distancia_km'].sum())} km")
            with col3:
                st.metric("Total Gastos", f"${formatear_numero(df_viajes['total_gastos'].sum())}")
            with col4:
                st.metric("Total Utilidad", f"${formatear_numero(df_viajes['utilidad'].sum())}")

            st.subheader("Resultados")

            columnas_mostrar = [
                'id', 'fecha_viaje', 'fecha_creacion', 'placa', 'conductor', 'origen', 'destino',
                'cliente', 'peso', 'distancia_km', 'dias_viaje', 'numero_viajes', 'total_gastos', 'valor_flete',
                'utilidad', 'rentabilidad'
            ]

            df_mostrar = df_viajes[columnas_mostrar].copy()
            df_mostrar.columns = [
                'ID', 'Fecha del Viaje', 'Fecha Registro', 'Placa', 'Conductor', 'Origen', 'Destino',
                'Cliente', 'Peso (kg)', 'Km', 'Días', 'N° Viajes', 'Total Gastos', 'Valor Flete', 'Utilidad', 'Rentabilidad %'
            ]

            df_mostrar['Total Gastos'] = df_mostrar['Total Gastos'].apply(lambda x: f"${formatear_numero(x)}")
            df_mostrar['Valor Flete'] = df_mostrar['Valor Flete'].apply(lambda x: f"${formatear_numero(x)}")
            df_mostrar['Utilidad'] = df_mostrar['Utilidad'].apply(lambda x: f"${formatear_numero(x)}")
            df_mostrar['Rentabilidad %'] = df_mostrar['Rentabilidad %'].apply(lambda x: f"{x:.1f}%")

            st.dataframe(df_mostrar, use_container_width=True, height=400)

            st.subheader("Ver Detalle de Viaje")

            def _etiqueta_viaje(vid):
                fila = df_viajes[df_viajes['id'] == vid].iloc[0]
                cliente_txt = f" | 🏢 {fila['cliente']}" if fila.get('cliente') else ""
                return f"#{vid} — {fila['fecha_viaje']} — {fila['placa']} — {fila['origen']} → {fila['destino']}{cliente_txt}"

            viaje_id_seleccionado = st.selectbox(
                "Selecciona un viaje",
                df_viajes['id'].tolist(),
                format_func=_etiqueta_viaje
            )

            if st.button("Ver Detalle Completo"):
                st.session_state.mostrar_detalle_viaje_id = viaje_id_seleccionado

            if st.session_state.get('mostrar_detalle_viaje_id') == viaje_id_seleccionado:
                viaje = db.obtener_viaje_por_id(viaje_id_seleccionado)
                if viaje:
                    col1, col2 = st.columns(2)

                    with col1:
                        st.markdown("### 📋 Información del Viaje")
                        st.write(f"**ID:** {viaje[0]}")
                        st.write(f"**Fecha de Registro:** {viaje[1]}")
                        try:
                            st.write(f"**📅 Fecha del Viaje:** {viaje[43]}")
                        except IndexError:
                            pass
                        st.write(f"**Placa:** {viaje[2]}")
                        try:
                            st.write(f"**🏢 Cliente:** {viaje[44] if viaje[44] else '-'}")
                        except IndexError:
                            pass
                        st.write(f"**Conductor:** {viaje[3]}")
                        st.write(f"**Ruta:** {viaje[4]} → {viaje[5]}")
                        st.write(f"**Distancia (efectiva del día):** {formatear_numero(viaje[6])} km")
                        st.write(f"**Días:** {viaje[7]}")
                        try:
                            st.write(f"**🚛 Número de Viajes:** {viaje[46] if viaje[46] else 1}")
                        except IndexError:
                            pass
                        try:
                            st.write(f"**⚖️ Peso:** {formatear_numero(viaje[47]) if len(viaje) > 47 and viaje[47] else 0} kg")
                        except IndexError:
                            pass
                        st.write(f"**Es Frontera:** {'Sí' if viaje[8] else 'No'}")
                        st.write(f"**Hubo Parqueo:** {'Sí' if viaje[9] else 'No'}")

                    with col2:
                        st.markdown("### 💰 Resultados Financieros")
                        st.write(f"**Total Gastos:** ${formatear_numero(viaje[28])}")
                        st.write(f"**Legalización:** ${formatear_numero(viaje[29])}")
                        st.write(f"**Punto Equilibrio:** ${formatear_numero(viaje[30])}")
                        st.write(f"**Valor Flete:** ${formatear_numero(viaje[31])}")

                        utilidad = viaje[32] if viaje[32] is not None else 0
                        rentabilidad = viaje[33] if viaje[33] is not None else 0

                        if utilidad >= 0:
                            st.success(f"**✅ Utilidad:** ${formatear_numero(utilidad)}")
                            st.success(f"**Rentabilidad:** {rentabilidad:.1f}%")
                        else:
                            st.error(f"**⚠️ Pérdida:** ${formatear_numero(utilidad)}")
                            st.error(f"**Rentabilidad:** {rentabilidad:.1f}%")

                        st.write(f"**Anticipo:** ${formatear_numero(viaje[34])}")
                        st.write(f"**Saldo:** ${formatear_numero(viaje[35])}")
                        st.write(f"**Hubo Anticipo Empresa:** {'Sí' if viaje[36] else 'No'}")
                        st.write(f"**Ant. Empresa (90%):** ${formatear_numero(viaje[37])}")
                        st.write(f"**Saldo Empresa:** ${formatear_numero(viaje[38])}")

                    st.markdown("### 📊 Desglose de Costos")
                    col1, col2, col3 = st.columns(3)

                    with col1:
                        st.write(f"1. Nómina Admin: ${formatear_numero(viaje[10])}")
                        st.write(f"2. Nómina Conductor: ${formatear_numero(viaje[11])}")
                        st.write(f"3. Comisión Conductor: ${formatear_numero(viaje[12])}")
                        st.write(f"4. Mantenimiento: ${formatear_numero(viaje[13])}")
                        st.write(f"5. Seguros: ${formatear_numero(viaje[14])}")
                        st.write(f"6. Tecnomecánica: ${formatear_numero(viaje[15])}")
                        st.write(f"7. Llantas: ${formatear_numero(viaje[16])}")

                    with col2:
                        st.write(f"8. Aceite: ${formatear_numero(viaje[17])}")
                        st.write(f"9. Combustible: ${formatear_numero(viaje[18])}")
                        st.write(f" - Galones: {formatear_decimal(viaje[19])}")
                        st.write(f"10. Flypass: ${formatear_numero(viaje[20])}")
                        st.write(f"11. Peajes: ${formatear_numero(viaje[21])}")
                        st.write(f"12. Cruce Frontera: ${formatear_numero(viaje[22])}")
                        st.write(f"13. Hotel: ${formatear_numero(viaje[23])}")

                    with col3:
                        st.write(f"14. Comida: ${formatear_numero(viaje[24])}")
                        st.write(f"15. Parqueo: ${formatear_numero(viaje[25])}")
                        st.write(f"16. Cargue/Descargue: ${formatear_numero(viaje[26])}")
                        st.write(f"17. Otros: ${formatear_numero(viaje[27])}")
                        try:
                            st.write(f"18. Urea/ACPM: ${formatear_numero(viaje[40])}")
                            st.write(f"19. Transporte: ${formatear_numero(viaje[41])}")
                            st.write(f"20. Propina/Comisión: ${formatear_numero(viaje[42])}")
                        except IndexError:
                            pass

                    if viaje[39]:
                        st.markdown("### 📝 Observaciones")
                        st.info(viaje[39])

                    col_ed, col_el = st.columns(2)
                    with col_ed:
                        if st.button("✏️ Editar este viaje", key=f"btn_editar_{viaje_id_seleccionado}"):
                            st.session_state.editando_viaje_id = viaje_id_seleccionado
                    with col_el:
                        if st.button("🗑️ Eliminar este viaje", type="secondary"):
                            db.eliminar_viaje(viaje_id_seleccionado)
                            st.success("Viaje eliminado")
                            st.rerun()

                    # ---------------- FORMULARIO DE EDICIÓN ----------------
                    if st.session_state.get('editando_viaje_id') == viaje_id_seleccionado:
                        st.divider()
                        st.markdown("### ✏️ Editar Viaje")
                        st.caption("Ajusta los valores necesarios. Los costos fijos (nómina, mantenimiento, seguros, etc.) se recalculan automáticamente con las fórmulas vigentes.")

                        placas_disponibles = [t.placa for t in st.session_state.tractomulas]
                        conductores_disponibles = [c.nombre for c in st.session_state.conductores]
                        rutas_disponibles = [f"{r.origen} → {r.destino}" for r in st.session_state.rutas]

                        placa_actual = viaje[2]
                        conductor_actual = viaje[3]
                        ruta_actual_str = f"{viaje[4]} → {viaje[5]}"

                        idx_placa = placas_disponibles.index(placa_actual) if placa_actual in placas_disponibles else 0
                        idx_conductor = conductores_disponibles.index(conductor_actual) if conductor_actual in conductores_disponibles else 0
                        idx_ruta = rutas_disponibles.index(ruta_actual_str) if ruta_actual_str in rutas_disponibles else 0

                        if not placas_disponibles or not conductores_disponibles or not rutas_disponibles:
                            st.error("⚠️ No se puede editar: faltan tractomulas, conductores o rutas registradas actualmente.")
                        else:
                            edit_ruta_str_preview = st.selectbox("Ruta", rutas_disponibles, index=idx_ruta, key="edit_ruta_preview")
                            edit_ruta_obj_preview = next(r for r in st.session_state.rutas if f"{r.origen} → {r.destino}" == edit_ruta_str_preview)
                            edit_numero_viajes_preview = st.number_input(
                                "🚛 Número de viajes", min_value=1,
                                value=int(viaje[46]) if len(viaje) > 46 and viaje[46] else 1,
                                step=1, key="edit_numero_viajes_preview",
                                help="Afecta la Comisión Conductor Urbano/Normal, y para AGOFER en rutas urbanas también el Flete, Cargue/Descargue y la distancia."
                            )
                            edit_cliente_preview = st.text_input(
                                "🏢 Cliente", value=(viaje[44] if len(viaje) > 44 and viaje[44] else ""), key="edit_cliente_preview"
                            )
                            edit_peso_preview_texto = st.text_input(
                                "⚖️ Peso transportado (kg)",
                                value=formatear_numero(viaje[47]) if len(viaje) > 47 and viaje[47] else "",
                                key="edit_peso_preview"
                            )
                            edit_peso_preview = limpiar_numero(edit_peso_preview_texto)

                            edit_aplica_agofer = edit_ruta_obj_preview.es_urbana and es_cliente_agofer(edit_cliente_preview)
                            edit_flete_sugerido = edit_peso_preview * datos.AGOFER_VALOR_POR_KG * edit_numero_viajes_preview if edit_aplica_agofer else 0.0
                            edit_cargue_sugerido = datos.AGOFER_CARGUE_DESCARGUE * edit_numero_viajes_preview if edit_aplica_agofer else 0.0

                            if edit_aplica_agofer:
                                st.success(
                                    f"🤖 Automatización AGOFER activa: Flete sugerido ${formatear_numero(edit_flete_sugerido)} · "
                                    f"Cargue/Descargue sugerido ${formatear_numero(edit_cargue_sugerido)} · "
                                    f"Distancia efectiva sugerida {formatear_numero(edit_ruta_obj_preview.distancia_km * edit_numero_viajes_preview)} km"
                                )

                            with st.form(key="form_editar_viaje"):
                                col1, col2 = st.columns(2)
                                with col1:
                                    edit_placa = st.selectbox("Tractomula", placas_disponibles, index=idx_placa, key="edit_placa")
                                    edit_conductor = st.selectbox("Conductor", conductores_disponibles, index=idx_conductor, key="edit_conductor")
                                    edit_dias = st.number_input("Días del viaje", min_value=1, value=int(viaje[7]), step=1, key="edit_dias")
                                    edit_fecha_viaje = st.date_input(
                                        "📅 Fecha del viaje",
                                        value=viaje[43] if len(viaje) > 43 and viaje[43] else datetime.now().date(),
                                        key="edit_fecha_viaje"
                                    )
                                with col2:
                                    edit_es_frontera = st.checkbox("¿Es viaje a frontera?", value=bool(viaje[8]), key="edit_frontera")
                                    edit_hubo_parqueo = st.checkbox("¿Hubo parqueo?", value=bool(viaje[9]), key="edit_parqueo")
                                    edit_hubo_ant_empresa = st.checkbox("¿Hubo anticipo empresa?", value=bool(viaje[36]), key="edit_ant_empresa")

                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    edit_flypass_texto = st.text_input("Flypass (COP)", value=formatear_numero(viaje[20]) if viaje[20] else "", key="edit_flypass")
                                    edit_flypass = limpiar_numero(edit_flypass_texto)
                                    edit_peajes_texto = st.text_input("Peajes (COP)", value=formatear_numero(viaje[21]) if viaje[21] else "", key="edit_peajes")
                                    edit_peajes = limpiar_numero(edit_peajes_texto)
                                    edit_urea_texto = st.text_input("Urea y/o ACPM (COP)", value=formatear_numero(viaje[40]) if len(viaje) > 40 and viaje[40] else "", key="edit_urea")
                                    edit_urea = limpiar_numero(edit_urea_texto)
                                with col2:
                                    edit_hotel_texto = st.text_input("Hotel (COP)", value=formatear_numero(viaje[23]) if viaje[23] else "", key="edit_hotel")
                                    edit_hotel = limpiar_numero(edit_hotel_texto)
                                    edit_comida_texto = st.text_input("Comida (COP)", value=formatear_numero(viaje[24]) if viaje[24] else "", key="edit_comida")
                                    edit_comida = limpiar_numero(edit_comida_texto)
                                    edit_transporte_texto = st.text_input("Transporte (COP)", value=formatear_numero(viaje[41]) if len(viaje) > 41 and viaje[41] else "", key="edit_transporte")
                                    edit_transporte = limpiar_numero(edit_transporte_texto)
                                with col3:
                                    edit_propina_texto = st.text_input("Propina/Comisión (COP)", value=formatear_numero(viaje[42]) if len(viaje) > 42 and viaje[42] else "", key="edit_propina")
                                    edit_propina = limpiar_numero(edit_propina_texto)
                                    valor_default_cargue_edit = edit_cargue_sugerido if edit_aplica_agofer else (viaje[26] or 0)
                                    edit_cargue_texto = st.text_input(
                                        "Cargue/Descargue (COP)",
                                        value=formatear_numero(valor_default_cargue_edit) if valor_default_cargue_edit else "",
                                        key="edit_cargue",
                                        help="Autocalculado para AGOFER en rutas urbanas: 30.000 x N° de Viajes. Editable."
                                    )
                                    edit_cargue = limpiar_numero(edit_cargue_texto)
                                    edit_otros_texto = st.text_input("Otros (COP)", value=formatear_numero(viaje[27]) if viaje[27] else "", key="edit_otros")
                                    edit_otros = limpiar_numero(edit_otros_texto)

                                col1, col2 = st.columns(2)
                                with col1:
                                    valor_default_flete_edit = edit_flete_sugerido if edit_aplica_agofer else (viaje[31] or 0)
                                    edit_valor_flete_texto = st.text_input(
                                        "Valor del Flete (COP)",
                                        value=formatear_numero(valor_default_flete_edit) if valor_default_flete_edit else "",
                                        key="edit_flete",
                                        help="Autocalculado para AGOFER en rutas urbanas: Peso x 27.500 x N° de Viajes. Editable."
                                    )
                                    edit_valor_flete = limpiar_numero(edit_valor_flete_texto)
                                with col2:
                                    edit_anticipo_texto = st.text_input("Anticipo (COP)", value=formatear_numero(viaje[34]) if viaje[34] else "", key="edit_anticipo")
                                    edit_anticipo = limpiar_numero(edit_anticipo_texto)

                                edit_observaciones = st.text_area("Observaciones", value=viaje[39] if viaje[39] else "", key="edit_obs")

                                col_guardar, col_cancelar = st.columns(2)
                                with col_guardar:
                                    guardar_edicion = st.form_submit_button("💾 Guardar Cambios", type="primary")
                                with col_cancelar:
                                    cancelar_edicion = st.form_submit_button("✖️ Cancelar Edición")

                                if guardar_edicion:
                                    edit_tractomula_obj = next(t for t in st.session_state.tractomulas if t.placa == edit_placa)
                                    edit_conductor_obj = next(c for c in st.session_state.conductores if c.nombre == edit_conductor)
                                    edit_ruta_obj = edit_ruta_obj_preview
                                    edit_numero_viajes = edit_numero_viajes_preview
                                    edit_cliente = edit_cliente_preview
                                    edit_peso = edit_peso_preview

                                    calculadora_editada = CalculadoraCostos(
                                        edit_tractomula_obj, edit_conductor_obj, edit_ruta_obj,
                                        edit_dias, edit_numero_viajes, edit_es_frontera, edit_hubo_parqueo,
                                        edit_flypass, edit_peajes, edit_urea, edit_hotel, edit_comida,
                                        edit_transporte, edit_propina, edit_cargue, edit_otros,
                                        edit_valor_flete, edit_anticipo, edit_hubo_ant_empresa, datos,
                                        peso=edit_peso, cliente=edit_cliente
                                    )
                                    exito = db.actualizar_viaje(viaje_id_seleccionado, calculadora_editada, edit_fecha_viaje, edit_observaciones, edit_cliente)
                                    if exito:
                                        st.success("✅ Viaje actualizado correctamente")
                                        del st.session_state.editando_viaje_id
                                        if 'ultima_busqueda' in st.session_state:
                                            del st.session_state.ultima_busqueda
                                        st.rerun()
                                if cancelar_edicion:
                                    del st.session_state.editando_viaje_id
                                    st.rerun()

            st.subheader("📥 Exportar Resultados")
            if st.button("Descargar en Excel"):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_viajes.to_excel(writer, sheet_name='Viajes', index=False)

                output.seek(0)
                fecha_descarga = datetime.now().strftime('%Y-%m-%d')
                st.download_button(
                    label="📥 Descargar Historial en Excel",
                    data=output,
                    file_name=f"Historial_Viajes_{fecha_descarga}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        st.subheader("📊 Estadísticas Generales")
        stats = db.obtener_estadisticas()

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("### Resumen Global")
            st.metric("Total de Viajes", stats['total_viajes'])
            st.metric("Total Kilómetros", f"{formatear_numero(stats['total_km'])} km")
            st.metric("Gastos Acumulados", f"${formatear_numero(stats['total_gastos'])}")

        with col2:
            st.markdown("### Top Tractomulas")
            if stats['viajes_por_placa']:
                for placa, total in stats['viajes_por_placa'][:5]:
                    st.write(f"**{placa}:** {total} viajes")
            else:
                st.info("No hay datos")

        with col3:
            st.markdown("### Top Conductores")
            if stats['viajes_por_conductor']:
                for conductor, total in stats['viajes_por_conductor'][:5]:
                    st.write(f"**{conductor}:** {total} viajes")
            else:
                st.info("No hay datos")

        if stats['rutas_frecuentes']:
            st.markdown("### Rutas Más Frecuentes")
            rutas_df = pd.DataFrame(stats['rutas_frecuentes'], columns=['Origen', 'Destino', 'Cantidad'])
            st.dataframe(rutas_df, use_container_width=True, hide_index=True)

        # ---------------- NUEVO v4.8: Días Sin Viaje (solo trazabilidad) ----------------
        st.divider()
        st.subheader("📭 Días Sin Viaje Registrados")
        st.caption("Registros de días en que la tractomula NO hizo viaje. No afectan ningún cálculo financiero, son solo para historial.")

        placa_f_vacio = st.selectbox("Filtrar por placa", ["Todas"] + sorted(PLACA_CONDUCTOR.keys()), key="vacio_filtro_placa")
        placa_f_vacio_val = None if placa_f_vacio == "Todas" else placa_f_vacio
        df_vacios = db.obtener_dias_sin_viaje(placa_f_vacio_val)

        if df_vacios.empty:
            st.info("No hay días sin viaje registrados.")
        else:
            st.metric("Total de Días Sin Viaje", len(df_vacios))
            df_mostrar_vacios = df_vacios[['id', 'fecha', 'placa', 'conductor', 'motivo', 'observaciones']].copy()
            df_mostrar_vacios.columns = ['ID', 'Fecha', 'Placa', 'Conductor', 'Motivo', 'Observaciones']
            st.dataframe(df_mostrar_vacios, use_container_width=True, hide_index=True)

            id_eliminar_vacio = st.selectbox(
                "Eliminar un registro",
                df_vacios['id'].tolist(),
                format_func=lambda vid: f"#{vid} — {df_vacios[df_vacios['id']==vid].iloc[0]['fecha']} — {df_vacios[df_vacios['id']==vid].iloc[0]['placa']}",
                key="vacio_id_eliminar"
            )
            if st.button("🗑️ Eliminar registro seleccionado", key="btn_eliminar_vacio"):
                db.eliminar_dia_sin_viaje(id_eliminar_vacio)
                st.success("Registro eliminado")
                st.rerun()

    # ==================== TAB 7: ACUMULADO POR FLOTA ====================
    if tab_actual == opciones_tabs[7]:
        st.header("Acumulado por Flota")
        st.markdown("Acumulados totales por unidad (tractomula/placa)")

        with st.expander("🔍 Filtros por Fecha", expanded=True):
            filtro_tipo = st.selectbox("Tipo de Filtro", ["Ninguno", "Mes", "Año", "Rango Personalizado"])
            fecha_inicio = None
            fecha_fin = None
            if filtro_tipo == "Mes":
                mes_seleccionado = st.selectbox("Mes", range(1, 13))
                año_seleccionado = st.selectbox("Año", range(2020, datetime.now().year + 1))
                fecha_inicio = f"{año_seleccionado}-{mes_seleccionado:02d}-01"
                if mes_seleccionado == 12:
                    fecha_fin = f"{año_seleccionado}-12-31"
                else:
                    ultimo_dia = (datetime(año_seleccionado, mes_seleccionado + 1, 1) - timedelta(days=1)).day
                    fecha_fin = f"{año_seleccionado}-{mes_seleccionado:02d}-{ultimo_dia}"
            elif filtro_tipo == "Año":
                año_seleccionado = st.selectbox("Año", range(2020, datetime.now().year + 1))
                fecha_inicio = f"{año_seleccionado}-01-01"
                fecha_fin = f"{año_seleccionado}-12-31"
            elif filtro_tipo == "Rango Personalizado":
                col1, col2 = st.columns(2)
                with col1:
                    fecha_inicio = st.date_input("Desde", value=None)
                with col2:
                    fecha_fin = st.date_input("Hasta", value=None)
                fecha_inicio = fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else None
                fecha_fin = fecha_fin.strftime('%Y-%m-%d') if fecha_fin else None

            buscar_totales = st.button("Aplicar Filtro", type="primary")

        primera_vez_totales = 'ultimos_totales' not in st.session_state
        if buscar_totales or primera_vez_totales:
            df_totales = db.obtener_totales_por_placa(fecha_inicio, fecha_fin)
            st.session_state.ultimos_totales = df_totales
        else:
            df_totales = st.session_state.ultimos_totales

        if df_totales.empty:
            st.info("No hay datos con los filtros aplicados.")
        else:
            df_mostrar = df_totales.copy()
            for col in df_mostrar.columns:
                if col != 'placa' and col.startswith('total_'):
                    if 'rentabilidad' in col:
                        df_mostrar[col] = df_mostrar[col].apply(lambda x: f"{x:.1f}%")
                    else:
                        df_mostrar[col] = df_mostrar[col].apply(lambda x: f"${formatear_numero(x)}")

            st.dataframe(df_mostrar, use_container_width=True)

            excel_totales = GeneradorReportes.generar_excel_totales(df_totales)
            st.download_button(
                label="📥 Descargar Reporte en Excel",
                data=excel_totales,
                file_name=f"Acumulados_Flota_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.subheader("Gráficos Comparativos")
            fig_cxc = px.bar(df_totales, x='placa', y='total_cxc', title="Total CXC por Unidad")
            st.plotly_chart(fig_cxc)
            fig_gastos = px.bar(df_totales, x='placa', y='total_gastos', title="Total Gastos por Unidad")
            st.plotly_chart(fig_gastos)
            fig_ut = px.bar(df_totales, x='placa', y='total_ut', title="Total UT por Unidad")
            st.plotly_chart(fig_ut)
            fig_rentabilidad = px.bar(df_totales, x='placa', y='total_rentabilidad', title="Rentabilidad por Unidad")
            st.plotly_chart(fig_rentabilidad)

    # ==================== TAB 8: LIQUIDACIONES DE CONDUCTORES ====================
    if tab_actual == opciones_tabs[8]:
        st.header("💵 Liquidaciones de Conductores")
        st.caption("Muestra cuántos viajes hizo el conductor, en qué placas, y el Total de Comisiones a pagar en el periodo (ej. quincena).")

        st.subheader("📝 Generar Nueva Liquidación")
        col1, col2, col3 = st.columns(3)
        with col1:
            conductores_nombres = [c.nombre for c in st.session_state.conductores] if st.session_state.conductores else []
            if conductores_nombres:
                conductor_liq = st.selectbox("Conductor", conductores_nombres, key="liq_conductor")
            else:
                conductor_liq = None
                st.warning("Primero registra conductores en la pestaña 3.")
        with col2:
            periodo_inicio_liq = st.date_input("Periodo Desde", value=datetime.now().replace(day=1).date(), key="liq_inicio")
        with col3:
            periodo_fin_liq = st.date_input("Periodo Hasta", value=datetime.now().date(), key="liq_fin")

        if conductor_liq and st.button("🔍 Buscar Viajes del Periodo", key="liq_buscar"):
            df_viajes_liq = db.obtener_viajes_para_liquidar(conductor_liq, periodo_inicio_liq, periodo_fin_liq)
            st.session_state.df_viajes_liq = df_viajes_liq
            st.session_state.conductor_liq_actual = conductor_liq
            st.session_state.periodo_liq_actual = (periodo_inicio_liq, periodo_fin_liq)

        if 'df_viajes_liq' in st.session_state and not st.session_state.df_viajes_liq.empty:
            df_viajes_liq = st.session_state.df_viajes_liq
            st.success(f"Se encontraron {len(df_viajes_liq)} viajes de {st.session_state.conductor_liq_actual} en el periodo seleccionado.")

            df_mostrar_liq = df_viajes_liq.copy()
            df_mostrar_liq['comision_conductor'] = df_mostrar_liq['comision_conductor'].apply(lambda x: f"${formatear_numero(x)}")
            df_mostrar_liq.columns = ['ID', 'Fecha Viaje', 'Placa', 'Origen', 'Destino', 'Comisión', 'N° Viajes']
            st.dataframe(df_mostrar_liq, use_container_width=True, hide_index=True)

            cantidad_viajes_preview = int(df_viajes_liq['numero_viajes'].fillna(1).sum())
            placas_preview = ", ".join(sorted(df_viajes_liq['placa'].unique()))
            total_comisiones_preview = float(df_viajes_liq['comision_conductor'].sum())

            col1, col2 = st.columns(2)
            with col1:
                st.metric("🚛 Cantidad de Viajes", cantidad_viajes_preview)
            with col2:
                st.metric("💰 TOTAL A PAGAR (Comisiones)", f"${formatear_numero(total_comisiones_preview)}")

            observaciones_liq = st.text_area("Observaciones de la liquidación (opcional)", key="liq_obs")

            if st.button("💾 Guardar Liquidación", type="primary", key="liq_guardar"):
                liquidacion_id, n_viajes, placas_liq, t_com, t_pagar = db.guardar_liquidacion(
                    st.session_state.conductor_liq_actual,
                    st.session_state.periodo_liq_actual[0],
                    st.session_state.periodo_liq_actual[1],
                    df_viajes_liq,
                    observaciones_liq
                )
                st.success(f"✅ Liquidación guardada (ID: {liquidacion_id}) — Total a pagar: ${formatear_numero(t_pagar)}")
                del st.session_state.df_viajes_liq
                st.rerun()
        elif 'df_viajes_liq' in st.session_state:
            st.info("No se encontraron viajes de este conductor en el periodo seleccionado.")

        st.divider()
        st.subheader("📋 Liquidaciones Registradas")

        col1, col2 = st.columns(2)
        with col1:
            filtro_conductor_liq = st.selectbox(
                "Filtrar por conductor",
                ["Todos"] + [c.nombre for c in st.session_state.conductores],
                key="filtro_liq_conductor"
            )
        with col2:
            filtro_estado_liq = st.selectbox("Filtrar por estado", ["Todos", "Pendiente", "Pagada"], key="filtro_liq_estado")

        cond_f = None if filtro_conductor_liq == "Todos" else filtro_conductor_liq
        est_f = None if filtro_estado_liq == "Todos" else filtro_estado_liq
        df_liquidaciones = db.obtener_liquidaciones(cond_f, est_f)

        if df_liquidaciones.empty:
            st.info("No hay liquidaciones registradas con estos filtros.")
        else:
            total_pendiente_liq = df_liquidaciones[df_liquidaciones['estado'] == 'Pendiente']['total_a_pagar'].sum()
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Liquidaciones", len(df_liquidaciones))
            with col2:
                st.metric("💰 Total Pendiente por Pagar", f"${formatear_numero(total_pendiente_liq)}")

            for _, liq in df_liquidaciones.iterrows():
                estado_icono = "🟢" if liq['estado'] == 'Pagada' else "🟡"
                with st.expander(f"{estado_icono} {liq['conductor']} | {liq['periodo_inicio']} → {liq['periodo_fin']} | ${formatear_numero(liq['total_a_pagar'])} | {liq['estado']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Conductor:** {liq['conductor']}")
                        st.write(f"**Periodo:** {liq['periodo_inicio']} a {liq['periodo_fin']}")
                        st.write(f"**🚛 Cantidad de Viajes:** {liq['cantidad_viajes']}")
                    with col2:
                        st.write(f"**💰 TOTAL A PAGAR (Comisiones):** ${formatear_numero(liq['total_a_pagar'])}")
                        st.write(f"**Estado:** {liq['estado']}" + (f" (pagada el {liq['fecha_pago']})" if liq['fecha_pago'] else ""))
                    if liq['observaciones']:
                        st.caption(f"📝 {liq['observaciones']}")

                    col_a, col_b = st.columns(2)
                    with col_a:
                        if liq['estado'] == 'Pendiente':
                            if st.button("✅ Marcar como Pagada", key=f"pagar_liq_{liq['id']}"):
                                db.marcar_liquidacion_pagada(liq['id'])
                                st.success("Liquidación marcada como pagada")
                                st.rerun()
                    with col_b:
                        if st.button("🗑️ Eliminar", key=f"eliminar_liq_{liq['id']}"):
                            db.eliminar_liquidacion(liq['id'])
                            st.success("Liquidación eliminada")
                            st.rerun()

    # ==================== TAB 9: CUENTAS PENDIENTES (POR PAGAR/COBRAR) ====================
    if tab_actual == opciones_tabs[9]:
        st.header("⏰ Pagos Pendientes y Vencimientos")
        st.caption("Controla lo que te deben (Por Cobrar: fletes de clientes) y lo que debes (Por Pagar: seguros, liquidaciones, proveedores, etc).")

        st.subheader("➕ Registrar Nueva Cuenta")
        with st.form(key="form_cuenta"):
            col1, col2 = st.columns(2)
            with col1:
                tipo_cuenta = st.selectbox("Tipo", ["Por Cobrar", "Por Pagar"])
                concepto_cuenta = st.text_input("Concepto", placeholder="Ej: Flete cliente XYZ, Seguro tractomula NOX459")
                tercero_cuenta = st.text_input("Tercero (cliente/proveedor)", placeholder="Nombre de la empresa o persona")
            with col2:
                monto_cuenta_texto = st.text_input("Monto (COP)", value="", placeholder="0")
                monto_cuenta = limpiar_numero(monto_cuenta_texto)
                if monto_cuenta > 0:
                    st.caption(f"💵 {formatear_numero(monto_cuenta)}")
                fecha_vencimiento_cuenta = st.date_input("Fecha de Vencimiento", value=datetime.now().date())
                observaciones_cuenta = st.text_area("Observaciones (opcional)")

            submit_cuenta = st.form_submit_button("💾 Guardar Cuenta", type="primary")
            if submit_cuenta:
                if not concepto_cuenta or monto_cuenta <= 0:
                    st.error("⚠️ Debes ingresar al menos el concepto y un monto mayor a cero.")
                else:
                    cuenta_id = db.guardar_cuenta(
                        tipo_cuenta, concepto_cuenta, tercero_cuenta, monto_cuenta,
                        fecha_vencimiento_cuenta, observaciones_cuenta
                    )
                    st.success(f"✅ Cuenta guardada (ID: {cuenta_id})")
                    st.rerun()

        st.divider()
        st.subheader("📋 Cuentas Registradas")

        col1, col2 = st.columns(2)
        with col1:
            filtro_tipo_cuenta = st.selectbox("Filtrar por tipo", ["Todas", "Por Cobrar", "Por Pagar"], key="filtro_cuenta_tipo")
        with col2:
            filtro_estado_cuenta = st.selectbox("Filtrar por estado", ["Todas", "Pendiente", "Pagado"], key="filtro_cuenta_estado")

        tipo_f = None if filtro_tipo_cuenta == "Todas" else filtro_tipo_cuenta
        estado_f = None if filtro_estado_cuenta == "Todas" else filtro_estado_cuenta
        df_cuentas = db.obtener_cuentas(tipo_f, estado_f)

        if df_cuentas.empty:
            st.info("No hay cuentas registradas con estos filtros.")
        else:
            hoy = datetime.now().date()
            df_pendientes = df_cuentas[df_cuentas['estado'] == 'Pendiente']
            total_por_cobrar = df_pendientes[df_pendientes['tipo'] == 'Por Cobrar']['monto'].sum()
            total_por_pagar = df_pendientes[df_pendientes['tipo'] == 'Por Pagar']['monto'].sum()

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("💚 Total Por Cobrar (pendiente)", f"${formatear_numero(total_por_cobrar)}")
            with col2:
                st.metric("🔴 Total Por Pagar (pendiente)", f"${formatear_numero(total_por_pagar)}")
            with col3:
                balance = total_por_cobrar - total_por_pagar
                st.metric("⚖️ Balance Neto", f"${formatear_numero(balance)}")

            st.divider()

            for _, cuenta in df_cuentas.iterrows():
                fecha_venc = cuenta['fecha_vencimiento']
                dias_para_vencer = (fecha_venc - hoy).days if cuenta['estado'] == 'Pendiente' else None

                if cuenta['estado'] == 'Pagado':
                    icono = "✅"
                elif dias_para_vencer is not None and dias_para_vencer < 0:
                    icono = "🔴"
                elif dias_para_vencer is not None and dias_para_vencer <= 7:
                    icono = "🟡"
                else:
                    icono = "🟢"

                tipo_icono = "💚" if cuenta['tipo'] == 'Por Cobrar' else "🔴"

                titulo = f"{icono} {tipo_icono} {cuenta['tipo']} | {cuenta['concepto']} | ${formatear_numero(cuenta['monto'])} | Vence: {fecha_venc}"
                with st.expander(titulo):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Tipo:** {cuenta['tipo']}")
                        st.write(f"**Concepto:** {cuenta['concepto']}")
                        st.write(f"**Tercero:** {cuenta['tercero'] or '-'}")
                        st.write(f"**Monto:** ${formatear_numero(cuenta['monto'])}")
                    with col2:
                        st.write(f"**Fecha de Vencimiento:** {fecha_venc}")
                        st.write(f"**Estado:** {cuenta['estado']}")
                        if cuenta['estado'] == 'Pendiente' and dias_para_vencer is not None:
                            if dias_para_vencer < 0:
                                st.error(f"⚠️ VENCIDA hace {abs(dias_para_vencer)} días")
                            elif dias_para_vencer <= 7:
                                st.warning(f"⏰ Vence en {dias_para_vencer} días")
                            else:
                                st.success(f"Vence en {dias_para_vencer} días")
                        if cuenta['fecha_pago']:
                            st.write(f"**Fecha de Pago:** {cuenta['fecha_pago']}")
                    if cuenta['observaciones']:
                        st.caption(f"📝 {cuenta['observaciones']}")

                    col_a, col_b = st.columns(2)
                    with col_a:
                        if cuenta['estado'] == 'Pendiente':
                            if st.button("✅ Marcar como Pagado", key=f"pagar_cuenta_{cuenta['id']}"):
                                db.marcar_cuenta_pagada(cuenta['id'])
                                st.success("Cuenta marcada como pagada")
                                st.rerun()
                    with col_b:
                        if st.button("🗑️ Eliminar", key=f"eliminar_cuenta_{cuenta['id']}"):
                            db.eliminar_cuenta(cuenta['id'])
                            st.success("Cuenta eliminada")
                            st.rerun()

    # ==================== TAB 10: SOBRECONSUMO DE COMBUSTIBLE ====================
    if tab_actual == opciones_tabs[10]:
        st.header("⛽ Detección de Sobreconsumo de Combustible")
        st.caption("Compara los galones que TEÓRICAMENTE debió gastar cada viaje (según distancia y rendimiento por tipo de ruta) contra los galones REALES que compraste. Solo aparecen aquí los viajes donde registraste el dato real.")

        st.info("💡 Para que un viaje aparezca acá, debes ingresar el campo **⛽ Galones Reales Comprados** al calcular el viaje (Tab 4) o al editarlo (Tab 6).")

        col1, col2 = st.columns(2)
        with col1:
            placa_filtro_consumo = st.selectbox("Filtrar por placa", ["Todas"] + sorted(PLACA_CONDUCTOR.keys()), key="filtro_consumo_placa")
        with col2:
            umbral_alerta = st.slider("Umbral de alerta de sobreconsumo (%)", min_value=5, max_value=50, value=10, step=5,
                                       help="Viajes con un % de sobreconsumo mayor a este valor se marcan en rojo")

        placa_f_consumo = None if placa_filtro_consumo == "Todas" else placa_filtro_consumo
        df_consumo = db.obtener_viajes_con_consumo(placa_f_consumo)

        if df_consumo.empty:
            st.info("No hay viajes con galones reales registrados todavía.")
        else:
            total_viajes_consumo = len(df_consumo)
            total_sobreconsumo = len(df_consumo[df_consumo['porcentaje_sobreconsumo'] > umbral_alerta])
            galones_extra_total = df_consumo[df_consumo['diferencia_galones'] > 0]['diferencia_galones'].sum()
            costo_extra_estimado = galones_extra_total * datos.PRECIO_DIESEL

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🚛 Viajes Analizados", total_viajes_consumo)
            with col2:
                st.metric("🔴 Viajes con Sobreconsumo", total_sobreconsumo)
            with col3:
                st.metric("💸 Costo Extra Estimado", f"${formatear_numero(costo_extra_estimado)}",
                          help="Galones de más x precio del diesel actual")

            st.divider()

            for _, v in df_consumo.iterrows():
                porcentaje = v['porcentaje_sobreconsumo']
                if porcentaje > umbral_alerta:
                    icono = "🔴"
                elif porcentaje > 0:
                    icono = "🟡"
                else:
                    icono = "🟢"

                titulo = (f"{icono} {v['placa']} | {v['conductor']} | {v['fecha_viaje']} | "
                          f"{v['origen']} → {v['destino']} | Sobreconsumo: {porcentaje:.1f}%")

                with st.expander(titulo):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.write(f"**Placa:** {v['placa']}")
                        st.write(f"**Conductor:** {v['conductor']}")
                        st.write(f"**Ruta:** {v['origen']} → {v['destino']} ({formatear_numero(v['distancia_km'])} km)")
                    with col2:
                        st.write(f"**Galones Teóricos:** {formatear_decimal(v['galones_necesarios'])} gal")
                        st.write(f"**Galones Reales:** {formatear_decimal(v['galones_reales'])} gal")
                        st.write(f"**Diferencia:** {formatear_decimal(v['diferencia_galones'])} gal")
                    with col3:
                        if porcentaje > umbral_alerta:
                            st.error(f"🔴 Sobreconsumo: {porcentaje:.1f}%")
                        elif porcentaje > 0:
                            st.warning(f"🟡 Leve diferencia: {porcentaje:.1f}%")
                        else:
                            st.success(f"🟢 Dentro de lo normal: {porcentaje:.1f}%")

                        if v['diferencia_galones'] > 0:
                            costo_extra_viaje = v['diferencia_galones'] * datos.PRECIO_DIESEL
                            st.caption(f"💸 Costo extra: ${formatear_numero(costo_extra_viaje)}")

                    if st.button("🗑️ Eliminar este viaje", key=f"eliminar_consumo_{v['id']}"):
                        db.eliminar_viaje(int(v['id']))
                        st.success("Viaje eliminado")
                        st.rerun()

            st.divider()
            st.subheader("📊 Tabla Completa")
            df_mostrar_consumo = df_consumo.copy()
            df_mostrar_consumo['galones_necesarios'] = df_mostrar_consumo['galones_necesarios'].apply(lambda x: formatear_decimal(x))
            df_mostrar_consumo['galones_reales'] = df_mostrar_consumo['galones_reales'].apply(lambda x: formatear_decimal(x))
            df_mostrar_consumo['diferencia_galones'] = df_mostrar_consumo['diferencia_galones'].apply(lambda x: formatear_decimal(x))
            df_mostrar_consumo['porcentaje_sobreconsumo'] = df_mostrar_consumo['porcentaje_sobreconsumo'].apply(lambda x: f"{x:.1f}%")
            df_mostrar_consumo = df_mostrar_consumo[['id', 'fecha_viaje', 'placa', 'conductor', 'origen', 'destino',
                                                       'galones_necesarios', 'galones_reales', 'diferencia_galones', 'porcentaje_sobreconsumo']]
            df_mostrar_consumo.columns = ['ID', 'Fecha', 'Placa', 'Conductor', 'Origen', 'Destino',
                                            'Gal. Teórico', 'Gal. Real', 'Diferencia', '% Sobreconsumo']
            st.dataframe(df_mostrar_consumo, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
